from __future__ import annotations

import io
import gzip
import json
import os
import tarfile
import threading
import zipfile
from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import Workbook

from projects.alpha_lab.data_spine.alpha_vantage import collect_alpha_vantage_free_proxies
from projects.alpha_lab.data_spine.boundary import build_boundary_attestation
from projects.alpha_lab.data_spine.bea_io import collect_bea_io_api, collect_bea_io_reference
from projects.alpha_lab.data_spine.cli import _load_local_env
from projects.alpha_lab.data_spine.eia_bulk import (
    collect_eia_large_bulk,
    materialize_eia_electricity_controls,
)
from projects.alpha_lab.data_spine.http import Response
from projects.alpha_lab.data_spine.http import decode_content
from projects.alpha_lab.data_spine.registry import load_registry
from projects.alpha_lab.data_spine.materialize import _next_session_open, materialize_earnings_events
from projects.alpha_lab.data_spine.market import materialize_market_panels
from projects.alpha_lab.data_spine.readiness import build_readiness
from projects.alpha_lab.data_spine.sec_bulk import (
    collect_sec_companyfacts_bulk,
    collect_sec_submissions_bulk,
)
from projects.alpha_lab.data_spine.sec_earnings import prepare_earnings_hydration_index
from projects.alpha_lab.data_spine.sec_delisting import (
    prepare_combined_8k_hydration_index,
    prepare_delisting_hydration_index,
)
from projects.alpha_lab.data_spine.sec_insiders import (
    _exclude_amendment_ambiguous_issuers,
    _parse_form4_original,
    materialize_original_insider_events,
)
from projects.alpha_lab.data_spine.sec_original_stream import capture_sec_original_stream
from projects.alpha_lab.data_spine.sources import (
    _parse_french_zip,
    audit_eia,
    audit_sharadar,
    collect_factors,
    collect_fred_alfred,
    collect_sec_reference,
    hydrate_sec_filings,
)
from projects.alpha_lab.data_spine.sharadar_bulk import capture_sharadar_bulk
from projects.alpha_lab.data_spine.sharadar_stream import capture_sharadar_stream
from projects.alpha_lab.data_spine.storage import output_root, write_bundle, write_bundle_from_paths
from projects.alpha_lab.data_spine.terminal_returns import (
    build_terminal_return_sensitivity,
)
from projects.alpha_lab.data_spine.usaspending import (
    _normalize_name,
    capture_usaspending_government_customer_proxy,
)
from projects.alpha_lab.data_spine.yfinance_analyst import collect_yfinance_analyst_proxy
from projects.alpha_lab.data_spine.vendor import validate_vendor_sample


UTC = timezone.utc
REGISTRY = load_registry()


def _response(body: bytes, headers=None) -> Response:
    return Response(body=body, status=200, headers=headers or {})


def test_http_content_decoding_supports_sec_gzip():
    payload = b'{"ok":true}'
    assert decode_content(gzip.compress(payload), {"content-encoding": "gzip"}) == payload


def test_next_session_open_skips_nyse_holidays_and_special_closures():
    assert _next_session_open(datetime(2026, 7, 2).date()) == "2026-07-06T13:30:00+00:00"
    assert _next_session_open(datetime(2012, 10, 26).date()) == "2012-10-31T13:30:00+00:00"
    assert _next_session_open(datetime(2018, 12, 4).date()) == "2018-12-06T14:30:00+00:00"


def test_registry_is_research_only_and_credential_names_only():
    assert REGISTRY.production_integration is False
    assert REGISTRY.output_root.startswith("outputs/research/alpha_lab/")
    text = json.dumps(REGISTRY.to_dict())
    assert "api_key_env" in text
    assert "api_key\":" not in text


def test_eia_large_bulk_streams_public_archive(tmp_path):
    def downloader(url, path):
        assert url.endswith("ELEC.zip")
        path.write_bytes(b"PK\x03\x04electricity")
        return {"last_modified": "today"}

    result = collect_eia_large_bulk(
        repo_root=tmp_path,
        registry=REGISTRY,
        dataset="electricity",
        downloader=downloader,
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    assert result["manifest"]["metadata"]["api_key_required"] is False
    assert result["paths"]["electricity.zip"].read_bytes() == b"PK\x03\x04electricity"


def test_bea_reference_is_industry_proxy_only(tmp_path):
    result = collect_bea_io_reference(
        repo_root=tmp_path,
        fetcher=lambda url: b"PDF" if url.endswith(".pdf") else b"XLSX",
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    assert result["manifest"]["metadata"]["not_issuer_relationship_edges"] is True
    assert len(result["paths"]) == 2


def test_bea_api_never_persists_key(tmp_path, monkeypatch):
    monkeypatch.setenv("BEA_API_KEY", "x" * 36)

    def fetcher(url):
        assert "x" * 36 in url
        return {"BEAAPI": {"Results": {"ParamValue": []}}}

    result = collect_bea_io_api(
        repo_root=tmp_path,
        fetcher=fetcher,
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    assert "x" * 36 not in json.dumps(result["manifest"])
    assert result["manifest"]["metadata"]["current_vintage_proxy_only"] is True


def test_alpha_vantage_proxy_is_bounded_and_not_pit(tmp_path, monkeypatch):
    monkeypatch.setenv("ALPHA_VANTAGE_API_KEY", "secret")

    def fetcher(url):
        assert "secret" in url
        if "LISTING_STATUS" in url:
            return b"symbol,name,exchange\nIBM,IBM,NYSE\n"
        return b'{"symbol":"IBM","estimates":[{"date":"2027-12-31"}]}'

    result = collect_alpha_vantage_free_proxies(
        repo_root=tmp_path,
        tickers=("IBM",),
        fetcher=fetcher,
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    manifest = result["manifest"]
    assert manifest["metadata"]["daily_request_count"] == 3
    assert manifest["metadata"]["not_historical_point_in_time"] is True
    assert "secret" not in json.dumps(manifest)


def test_alpha_vantage_proxy_rejects_over_free_daily_budget(tmp_path, monkeypatch):
    monkeypatch.setenv("ALPHA_VANTAGE_API_KEY", "secret")
    with pytest.raises(ValueError, match="between 0 and 23"):
        collect_alpha_vantage_free_proxies(repo_root=tmp_path, max_tickers=24)


def test_delisting_hydration_index_is_candidate_only(tmp_path):
    master = tmp_path / "data/pit_universe/security_master.csv"
    master.parent.mkdir(parents=True)
    master.write_text(
        "security_id,permaticker,cik,cusip,figi,ticker,name,exchange,category,sector,industry,effective_start,effective_end,firstpricedate,lastpricedate,relatedtickers,source,confidence\n"
        "SEC-1,1,1,,,ABC,ACME,NYSE,Domestic Common Stock,Industrials,,2010-01-01,2020-01-15,,,,,,\n",
        encoding="utf-8",
    )
    write_bundle(
        repo_root=tmp_path,
        source_id="sharadar_actions",
        files={"actions.jsonl": b'{"action":"acquisitionby","date":"2020-01-15","ticker":"ABC","contraticker":"XYZ"}\n'},
        metadata={"test": True},
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    write_bundle(
        repo_root=tmp_path,
        source_id="sec_event_index",
        files={
            "event_index.csv": (
                b"cik,company_name,form_type,filed_date,filename,index_year,index_quarter\n"
                b"1,ACME,8-K,2020-01-16,edgar/data/1/a.txt,2020,1\n"
            ),
            "errors.json": b"[]\n",
        },
        metadata={"test": True},
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    result = prepare_delisting_hydration_index(tmp_path)
    assert result["delisting_hydration_candidate_rows"] == 1
    manifest = json.loads(Path(result["delisting_candidate_manifest_path"]).read_text())
    assert manifest["settlement_certified"] is False
    assert "terminal_settlement_not_parsed" in " ".join(manifest["blockers"])


def test_combined_8k_index_deduplicates_filenames(tmp_path):
    shared = tmp_path / "outputs/research/alpha_lab/shared"
    shared.mkdir(parents=True)
    header = "cik,company_name,form_type,filed_date,filename,index_year,index_quarter\n"
    row = "1,ACME,8-K,2020-01-16,edgar/data/1/a.txt,2020,1\n"
    (shared / "earnings_8k_hydration_index.csv").write_text(header + row, encoding="utf-8")
    (shared / "delisting_8k_hydration_index.csv").write_text(header + row, encoding="utf-8")
    result = prepare_combined_8k_hydration_index(tmp_path)
    assert result["combined_8k_hydration_candidate_rows"] == 1


def test_eia_electricity_controls_are_labeled_current_vintage(tmp_path):
    archive_path = tmp_path / "electricity.zip"
    record = {
        "series_id": "ELEC.PRICE.US-IND.M",
        "name": "Industrial price",
        "units": "cents",
        "f": "M",
        "last_updated": "2026-06-25T02:00:00-04:00",
        "data": [["202606", 9.0], ["201001", 6.0]],
    }
    with zipfile.ZipFile(archive_path, "w") as archive:
        archive.writestr("ELEC.txt", json.dumps(record) + "\n")
    write_bundle_from_paths(
        repo_root=tmp_path,
        source_id="eia_electricity_bulk",
        files={"electricity.zip": archive_path},
        metadata={"test": True},
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    result = materialize_eia_electricity_controls(tmp_path)
    assert result["eia_electricity_control_rows"] == 1
    text = Path(result["eia_electricity_controls_path"]).read_text()
    assert "CURRENT_BULK_VINTAGE_PROXY_ONLY" in text
    assert "201001" not in text


def test_usaspending_proxy_accepts_only_exact_normalized_issuer(tmp_path):
    master = tmp_path / "data/pit_universe/security_master.csv"
    master.parent.mkdir(parents=True)
    master.write_text(
        "security_id,permaticker,cik,cusip,figi,ticker,name,exchange,category,sector,industry,effective_start,effective_end,firstpricedate,lastpricedate,relatedtickers,source,confidence\n"
        "SEC-1,1,1,,,ABC,ACME CORPORATION,NYSE,Domestic Common Stock,Industrials,,2010-01-01,,,,,,\n",
        encoding="utf-8",
    )

    def poster(_url, payload):
        if "autocomplete/recipient" in _url:
            assert payload["search_text"] == "ACME CORPORATION"
            return {
                "results": [
                    {"recipient_name": "ACME CORP", "uei": "UEI1"},
                    {"recipient_name": "ACME SERVICES LLC", "uei": "UEI2"},
                ]
            }
        assert payload["filters"]["recipient_search_text"] == ["UEI1"]
        return {
            "results": [
                {"Award ID": "A1", "Recipient Name": "ACME CORP", "Recipient UEI": "UEI1", "Start Date": "2020-01-02", "End Date": "2021-01-02", "Award Amount": 10, "Awarding Agency": "Agency"},
                {"Award ID": "B1", "Recipient Name": "ACME SERVICES LLC", "Recipient UEI": "UEI2", "Start Date": "2020-01-02", "End Date": "2021-01-02", "Award Amount": 20, "Awarding Agency": "Agency"},
            ],
            "page_metadata": {"hasNext": False},
        }

    result = capture_usaspending_government_customer_proxy(
        repo_root=tmp_path,
        partition_size=1,
        poster=poster,
        sleeper=lambda _seconds: None,
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    assert _normalize_name("Acme Corporation") == "ACME CORP"
    assert result["manifest"]["metadata"]["government_customer_edge_count"] == 1
    data_path = next(path for name, path in result["paths"].items() if name.startswith("edges/"))
    with gzip.open(data_path, "rt", encoding="utf-8") as stream:
        edge = json.loads(next(stream))
    assert edge["supplier_security_id"] == "SEC-1"
    assert edge["recipient_uei"] == "UEI1"


def test_yfinance_analyst_proxy_is_forward_only(tmp_path):
    result = collect_yfinance_analyst_proxy(
        repo_root=tmp_path,
        tickers=("IBM",),
        fetcher=lambda symbol: {"symbol": symbol, "eps_revisions": {"0y": 1}},
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    metadata = result["manifest"]["metadata"]
    assert metadata["captured_ticker_count"] == 1
    assert metadata["usable_ticker_count"] == 1
    assert metadata["not_historical_point_in_time"] is True
    assert metadata["no_analyst_or_broker_identity"] is True


def test_cli_loads_ignored_local_env_without_overriding_caller(tmp_path, monkeypatch):
    (tmp_path / ".env").write_text(
        "NASDAQ_DATA_LINK_API_KEY='local-value'\nSEC_USER_AGENT=Research test@example.com\n",
        encoding="utf-8",
    )
    monkeypatch.setenv("NASDAQ_DATA_LINK_API_KEY", "caller-value")
    monkeypatch.delenv("SEC_USER_AGENT", raising=False)
    _load_local_env(tmp_path)
    assert os.environ["NASDAQ_DATA_LINK_API_KEY"] == "caller-value"
    assert os.environ["SEC_USER_AGENT"] == "Research test@example.com"


def test_bundle_is_immutable_and_confined(tmp_path):
    timestamp = datetime(2026, 7, 15, 12, 0, tzinfo=UTC)
    first = write_bundle(
        repo_root=tmp_path,
        source_id="test_source",
        files={"raw/data.txt": b"value\n"},
        metadata={"test": True},
        retrieved_at=timestamp,
    )
    second = write_bundle(
        repo_root=tmp_path,
        source_id="test_source",
        files={"raw/data.txt": b"value\n"},
        metadata={"test": True},
        retrieved_at=timestamp,
    )
    assert first["bundle_id"] == second["bundle_id"]
    assert first["manifest_path"].is_relative_to(output_root(tmp_path))
    changed = write_bundle(
        repo_root=tmp_path,
        source_id="test_source",
        files={"raw/data.txt": b"different\n"},
        metadata={"test": True},
        retrieved_at=timestamp,
    )
    assert changed["bundle_id"] != first["bundle_id"]


def test_streaming_bundle_and_bulk_export_do_not_persist_credentials(tmp_path, monkeypatch):
    source = tmp_path / "source.zip"
    source.write_bytes(b"PK\x03\x04test")
    streamed = write_bundle_from_paths(
        repo_root=tmp_path,
        source_id="streamed",
        files={"sample.zip": source},
        metadata={"test": True},
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    assert streamed["paths"]["sample.zip"].read_bytes() == source.read_bytes()

    monkeypatch.setenv("NASDAQ_DATA_LINK_API_KEY", "secret-value")
    metadata = {
        "datatable_bulk_download": {
            "file": {
                "status": "fresh",
                "link": "https://example.test/file.zip?api_key=secret-value",
                "data_snapshot_time": "2026-07-15T00:00:00Z",
            }
        }
    }

    def downloader(url, path):
        assert "secret-value" in url
        path.write_bytes(b"PK\x03\x04bulk")

    result = capture_sharadar_bulk(
        repo_root=tmp_path,
        registry=REGISTRY,
        table="DAILY",
        columns=("ticker", "date", "marketcap"),
        fetcher=lambda *args, **kwargs: _response(json.dumps(metadata).encode()),
        downloader=downloader,
        sleeper=lambda _seconds: None,
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    manifest_text = result["manifest_path"].read_text()
    assert "secret-value" not in manifest_text
    assert "example.test" not in manifest_text


def test_paginated_sharadar_stream_is_compressed_and_complete(tmp_path, monkeypatch):
    monkeypatch.setenv("NASDAQ_DATA_LINK_API_KEY", "secret-value")
    pages = iter(
        (
            {
                "datatable": {
                    "columns": [{"name": "ticker"}, {"name": "date"}, {"name": "marketcap"}],
                    "data": [["A", "2026-01-02", 10]],
                },
                "meta": {"next_cursor_id": "next"},
            },
            {
                "datatable": {
                    "columns": [{"name": "ticker"}, {"name": "date"}, {"name": "marketcap"}],
                    "data": [["B", "2026-01-03", 20]],
                },
                "meta": {"next_cursor_id": None},
            },
        )
    )
    result = capture_sharadar_stream(
        repo_root=tmp_path,
        registry=REGISTRY,
        table="DAILY",
        columns=("ticker", "date", "marketcap"),
        fetcher=lambda *args, **kwargs: _response(json.dumps(next(pages)).encode()),
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    with gzip.open(result["paths"]["daily.csv.gz"], "rt", encoding="utf-8") as stream:
        assert stream.read().splitlines() == [
            "ticker,date,marketcap",
            "A,2026-01-02,10",
            "B,2026-01-03,20",
        ]
    assert result["manifest"]["metadata"]["row_count"] == 2
    assert result["manifest"]["metadata"]["pagination_complete"] is True


def test_sharadar_stream_retries_and_resumes_completed_ticker_chunks(tmp_path, monkeypatch):
    monkeypatch.setenv("NASDAQ_DATA_LINK_API_KEY", "secret-value")
    columns = [{"name": "ticker"}, {"name": "date"}, {"name": "marketcap"}]
    first_calls = []

    def first_fetch(url, **kwargs):
        first_calls.append(url)
        if "ticker=A" in url:
            return _response(
                json.dumps(
                    {
                        "datatable": {"columns": columns, "data": [["A", "2026-01-02", 10]]},
                        "meta": {"next_cursor_id": None},
                    }
                ).encode()
            )
        raise TimeoutError("transient")

    with pytest.raises(RuntimeError, match="bounded retries: TimeoutError"):
        capture_sharadar_stream(
            repo_root=tmp_path,
            registry=REGISTRY,
            table="DAILY",
            columns=("ticker", "date", "marketcap"),
            tickers=("A", "B"),
            ticker_chunk_size=1,
            fetcher=first_fetch,
            request_attempts=1,
        )
    second_calls = []

    def second_fetch(url, **kwargs):
        second_calls.append(url)
        assert "ticker=A" not in url
        return _response(
            json.dumps(
                {
                    "datatable": {"columns": columns, "data": [["B", "2026-01-03", 20]]},
                    "meta": {"next_cursor_id": None},
                }
            ).encode()
        )

    result = capture_sharadar_stream(
        repo_root=tmp_path,
        registry=REGISTRY,
        table="DAILY",
        columns=("ticker", "date", "marketcap"),
        tickers=("A", "B"),
        ticker_chunk_size=1,
        fetcher=second_fetch,
        request_attempts=1,
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    with gzip.open(result["paths"]["daily.csv.gz"], "rt", encoding="utf-8") as stream:
        assert stream.read().splitlines() == [
            "ticker,date,marketcap",
            "A,2026-01-02,10",
            "B,2026-01-03,20",
        ]
    assert len(first_calls) == 2
    assert len(second_calls) == 1


def test_sharadar_audit_blocks_without_credential(tmp_path, monkeypatch):
    monkeypatch.delenv("NASDAQ_DATA_LINK_API_KEY", raising=False)
    result = audit_sharadar(
        repo_root=tmp_path,
        registry=REGISTRY,
        checked_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    assert result["audit"]["credential_present"] is False
    assert all(row["status"] == "BLOCKED_CREDENTIAL" for row in result["audit"]["tables"])


def test_sec_reference_requires_contact_and_never_persists_user_agent(tmp_path, monkeypatch):
    monkeypatch.delenv("SEC_USER_AGENT", raising=False)
    with pytest.raises(RuntimeError, match="contact email"):
        collect_sec_reference(repo_root=tmp_path, registry=REGISTRY)
    payload = {"fields": ["cik", "name", "ticker", "exchange"], "data": [[1, "A", "A", "NYSE"]]}
    result = collect_sec_reference(
        repo_root=tmp_path,
        registry=REGISTRY,
        user_agent="Caerus Research test@example.com",
        fetcher=lambda *args, **kwargs: _response(json.dumps(payload).encode()),
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    manifest = result["manifest"]
    assert manifest["metadata"]["row_count"] == 1
    assert "test@example.com" not in json.dumps(manifest)


def test_sec_hydration_preserves_original_and_converts_acceptance_time(tmp_path):
    index = tmp_path / "index.csv"
    index.write_text(
        "cik,company_name,form_type,filed_date,filename,index_year,index_quarter\n"
        "1,Issuer,4,2026-07-14,edgar/data/1/a.txt,2026,3\n",
        encoding="utf-8",
    )
    body = (
        b"<SEC-HEADER>\nACCESSION NUMBER: 0000000001-26-000001\n"
        b"ACCEPTANCE-DATETIME: 20260714173000\n</SEC-HEADER>\n"
    )
    result = hydrate_sec_filings(
        repo_root=tmp_path,
        registry=REGISTRY,
        index_path=index,
        limit=1,
        user_agent="Caerus Research test@example.com",
        fetcher=lambda *args, **kwargs: _response(body),
        sleeper=lambda _seconds: None,
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    inventory = result["paths"]["filing_inventory.csv"].read_text()
    assert "2026-07-14T21:30:00+00:00" in inventory
    assert result["manifest"]["metadata"]["original_submission_preserved"] is True


def test_sec_hydration_parses_live_submission_header_acceptance_form(tmp_path):
    index = tmp_path / "index.csv"
    index.write_text(
        "cik,company_name,form_type,filed_date,filename,index_year,index_quarter\n"
        "1,Issuer,4,2026-07-14,edgar/data/1/a.txt,2026,3\n",
        encoding="utf-8",
    )
    body = (
        b"<SEC-DOCUMENT>\n<ACCEPTANCE-DATETIME>20260714173000\n"
        b"ACCESSION NUMBER:\t0000000001-26-000001\n"
    )
    result = hydrate_sec_filings(
        repo_root=tmp_path,
        registry=REGISTRY,
        index_path=index,
        limit=1,
        user_agent="Caerus Research test@example.com",
        fetcher=lambda *args, **kwargs: _response(body),
        sleeper=lambda _seconds: None,
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    inventory = result["paths"]["filing_inventory.csv"].read_text()
    assert "2026-07-14T21:30:00+00:00" in inventory


def test_sec_original_stream_partitions_and_preserves_filings(tmp_path):
    index = tmp_path / "index.csv"
    index.write_text(
        "cik,company_name,form_type,filed_date,filename,index_year,index_quarter\n"
        "1,Issuer,4,2026-07-14,edgar/data/1/a.txt,2026,3\n"
        "2,Issuer,4/A,2026-07-15,edgar/data/2/b.txt,2026,3\n",
        encoding="utf-8",
    )

    calls = []

    def fetcher(url, **_kwargs):
        calls.append(url)
        accession = "0000000001-26-000001" if url.endswith("a.txt") else "0000000002-26-000002"
        body = (
            "<SEC-HEADER>\nACCESSION NUMBER: {}\n"
            "ACCEPTANCE-DATETIME: 20260714173000\n</SEC-HEADER>\n"
        ).format(accession).encode()
        return _response(body)

    progress = capture_sec_original_stream(
        repo_root=tmp_path,
        registry=REGISTRY,
        index_path=index,
        partition_size=1,
        max_new_partitions=1,
        user_agent="Caerus Research test@example.com",
        fetcher=fetcher,
        sleeper=lambda _seconds: None,
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    assert progress["capture_status"] == "IN_PROGRESS"
    assert progress["completed_partition_count"] == 1
    result = capture_sec_original_stream(
        repo_root=tmp_path,
        registry=REGISTRY,
        index_path=index,
        partition_size=1,
        user_agent="Caerus Research test@example.com",
        fetcher=fetcher,
        sleeper=lambda _seconds: None,
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    metadata = result["manifest"]["metadata"]
    assert metadata["candidate_count"] == 2
    assert metadata["partition_count"] == 2
    assert metadata["acceptance_timestamp_pass_count"] == 2
    assert len([name for name in result["paths"] if name.startswith("partitions/")]) == 2
    assert len(calls) == 2


def test_sec_original_stream_bounds_in_flight_response_bodies(tmp_path):
    index = tmp_path / "index.csv"
    rows = [
        "{cik},Issuer,8-K,2026-07-{day:02d},edgar/data/{cik}/{cik}.txt,2026,3".format(
            cik=cik,
            day=((cik - 1) % 9) + 1,
        )
        for cik in range(1, 13)
    ]
    index.write_text(
        "cik,company_name,form_type,filed_date,filename,index_year,index_quarter\n"
        + "\n".join(reversed(rows))
        + "\n",
        encoding="utf-8",
    )

    lock = threading.Lock()
    live_bodies = 0
    peak_live_bodies = 0

    class TrackedBody(bytes):
        def __new__(cls, payload):
            nonlocal live_bodies, peak_live_bodies
            instance = super().__new__(cls, payload)
            with lock:
                live_bodies += 1
                peak_live_bodies = max(peak_live_bodies, live_bodies)
            return instance

        def __del__(self):
            nonlocal live_bodies
            with lock:
                live_bodies -= 1

    def fetcher(url, **_kwargs):
        cik = Path(url).stem
        body = TrackedBody(
            (
                "<SEC-HEADER>\nACCESSION NUMBER: 000000{cik:0>4}-26-000001\n"
                "ACCEPTANCE-DATETIME: 20260714173000\n</SEC-HEADER>\n"
            ).format(cik=cik).encode()
            + (b"x" * (256 * 1024))
        )
        return _response(body)

    result = capture_sec_original_stream(
        repo_root=tmp_path,
        registry=REGISTRY,
        index_path=index,
        forms=("8-K",),
        partition_size=12,
        request_workers=2,
        user_agent="Caerus Research test@example.com",
        fetcher=fetcher,
        sleeper=lambda _seconds: None,
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )

    assert result["manifest"]["metadata"]["hydrated_count"] == 12
    assert peak_live_bodies <= 2
    assert live_bodies == 0


def test_sec_original_stream_disk_index_preserves_canonical_order(tmp_path):
    index = tmp_path / "index.csv"
    index.write_text(
        "cik,company_name,form_type,filed_date,filename,index_year,index_quarter\n"
        "9,Issuer,8-K,2026-07-15,edgar/data/9/z.txt,2026,3\n"
        "2,Issuer,8-K,2026-07-14,edgar/data/2/b.txt,2026,3\n"
        "1,Issuer,8-K,2026-07-14,edgar/data/1/a.txt,2026,3\n",
        encoding="utf-8",
    )

    def fetcher(url, **_kwargs):
        cik = url.split("/edgar/data/", 1)[1].split("/", 1)[0]
        accession = "000000{:0>4}-26-000001".format(cik)
        body = (
            "<SEC-HEADER>\nACCESSION NUMBER: {}\n"
            "ACCEPTANCE-DATETIME: 20260714173000\n</SEC-HEADER>\n"
        ).format(accession).encode()
        return _response(body)

    result = capture_sec_original_stream(
        repo_root=tmp_path,
        registry=REGISTRY,
        index_path=index,
        forms=("8-K",),
        partition_size=3,
        request_workers=2,
        user_agent="Caerus Research test@example.com",
        fetcher=fetcher,
        sleeper=lambda _seconds: None,
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    archive_path = result["paths"]["partitions/part_00000.tar.gz"]
    with tarfile.open(archive_path, "r:gz") as archive:
        assert archive.getnames() == [
            "filings/0000000001-26-000001.txt",
            "filings/0000000002-26-000001.txt",
            "filings/0000000009-26-000001.txt",
        ]


def test_sec_companyfacts_bulk_streams_without_persisting_user_agent(tmp_path):
    def downloader(url, path, user_agent):
        assert url.endswith("companyfacts.zip")
        assert user_agent == "Caerus Research test@example.com"
        path.write_bytes(b"PK\x03\x04facts")
        return {"last_modified": "today"}

    result = collect_sec_companyfacts_bulk(
        repo_root=tmp_path,
        registry=REGISTRY,
        user_agent="Caerus Research test@example.com",
        downloader=downloader,
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    manifest_text = result["manifest_path"].read_text()
    assert "test@example.com" not in manifest_text
    assert result["paths"]["companyfacts.zip"].read_bytes() == b"PK\x03\x04facts"


def test_sec_submissions_bulk_streams_without_persisting_user_agent(tmp_path):
    def downloader(url, path, user_agent):
        assert url.endswith("submissions.zip")
        assert user_agent == "Caerus Research test@example.com"
        path.write_bytes(b"PK\x03\x04submissions")
        return {"last_modified": "today"}

    result = collect_sec_submissions_bulk(
        repo_root=tmp_path,
        registry=REGISTRY,
        user_agent="Caerus Research test@example.com",
        downloader=downloader,
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    assert "test@example.com" not in result["manifest_path"].read_text()
    assert result["paths"]["submissions.zip"].read_bytes() == b"PK\x03\x04submissions"


def test_materialize_earnings_events_uses_item_202_and_exact_acceptance(tmp_path):
    archive_path = tmp_path / "submissions.zip"
    payload = {
        "cik": "1",
        "filings": {
            "recent": {
                "accessionNumber": ["0000000001-26-000001", "0000000001-26-000002"],
                "filingDate": ["2026-06-29", "2026-06-29"],
                "reportDate": ["2026-06-30", "2026-06-30"],
                "acceptanceDateTime": ["2026-06-29T17:30:00.000Z", "2026-06-29T18:00:00.000Z"],
                "form": ["8-K", "8-K"],
                "items": ["2.02,9.01", "1.01"],
            }
        },
    }
    with zipfile.ZipFile(archive_path, "w") as archive:
        archive.writestr("CIK0000000001.json", json.dumps(payload))
    write_bundle_from_paths(
        repo_root=tmp_path,
        source_id="sec_submissions",
        files={"submissions.zip": archive_path},
        metadata={"test": True},
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    master = tmp_path / "data/pit_universe/security_master.csv"
    master.parent.mkdir(parents=True)
    master.write_text(
        "security_id,cik,category,effective_start,effective_end\n"
        "SEC-1,1,Domestic Common Stock,2010-01-01,\n",
        encoding="utf-8",
    )
    facts = tmp_path / "outputs/research/alpha_lab/shared/sec_companyfacts_compact.csv.gz"
    facts.parent.mkdir(parents=True)
    with gzip.open(facts, "wt", encoding="utf-8", newline="") as stream:
        stream.write(
            "cik,entity_name,logical_fact,taxonomy,source_fact,unit,value,start,end,filed,available_at,accession_number,form,fiscal_year,fiscal_period,frame\n"
            "1,Issuer,eps_diluted,us-gaap,EarningsPerShareDiluted,USD/shares,1.25,2026-04-01,2026-06-30,2026-07-14,,0000000001-26-000001,8-K,2026,Q2,\n"
        )
    result = materialize_earnings_events(tmp_path)
    assert result["earnings_event_rows"] == 1
    with gzip.open(result["earnings_event_path"], "rt", encoding="utf-8") as stream:
        event = json.loads(next(stream))
    assert event["security_id"] == "SEC-1"
    assert event["acceptance_datetime_utc"] == "2026-06-29T17:30:00+00:00"
    assert event["available_at"] == "2026-06-30T13:30:00+00:00"
    assert event["announcement_time"] is None
    assert event["source_sha256"] is None
    assert event["reported_eps"] == "1.25"
    index_result = prepare_earnings_hydration_index(tmp_path)
    assert index_result["earnings_hydration_candidate_rows"] == 1
    index_text = Path(index_result["earnings_hydration_index_path"]).read_text()
    assert "edgar/data/1/000000000126000001/0000000001-26-000001.txt" in index_text


def test_original_form4_parser_accepts_sec_numeric_relationship_flags():
    parsed = _parse_form4_original(
        """<ownershipDocument>
        <issuer><issuerCik>123</issuerCik></issuer>
        <reportingOwner>
          <reportingOwnerId><rptOwnerCik>456</rptOwnerCik></reportingOwnerId>
          <reportingOwnerRelationship><isDirector>1</isDirector><isOfficer>0</isOfficer><isTenPercentOwner>true</isTenPercentOwner></reportingOwnerRelationship>
        </reportingOwner>
        <nonDerivativeTable><nonDerivativeTransaction>
          <transactionDate><value>2026-07-14</value></transactionDate>
          <transactionCoding><transactionCode>P</transactionCode></transactionCoding>
          <transactionAmounts><transactionShares><value>10</value></transactionShares><transactionPricePerShare><value>2.5</value></transactionPricePerShare></transactionAmounts>
        </nonDerivativeTransaction></nonDerivativeTable>
        </ownershipDocument>"""
    )
    assert parsed["issuer_cik"] == "0000000123"
    owner = parsed["owners"][0]
    assert owner["cik"] == "456"
    assert owner["is_director"] is True
    assert owner["is_officer"] is False
    assert owner["is_ten_percent_owner"] is True
    assert parsed["purchase_value"] == 25.0


def test_original_form4_materialization_uses_xml_as_canonical_source(tmp_path):
    index = tmp_path / "index.csv"
    index.write_text(
        "cik,company_name,form_type,filed_date,filename,index_year,index_quarter\n"
        "123,Issuer,4,2026-07-14,edgar/data/123/a.txt,2026,3\n",
        encoding="utf-8",
    )
    body = b"""<SEC-DOCUMENT>
<SEC-HEADER>
ACCESSION NUMBER: 0000000123-26-000001
ACCEPTANCE-DATETIME: 20260714173000
</SEC-HEADER>
<XML>
<ownershipDocument>
  <documentType>4</documentType>
  <periodOfReport>2026-07-14</periodOfReport>
  <issuer><issuerCik>123</issuerCik><issuerName>Issuer</issuerName><issuerTradingSymbol>TST</issuerTradingSymbol></issuer>
  <reportingOwner>
    <reportingOwnerId><rptOwnerCik>456</rptOwnerCik><rptOwnerName>TEST PERSON</rptOwnerName></reportingOwnerId>
    <reportingOwnerRelationship><isDirector>1</isDirector><isOfficer>1</isOfficer><isTenPercentOwner>0</isTenPercentOwner><officerTitle>Chief Executive Officer</officerTitle></reportingOwnerRelationship>
  </reportingOwner>
  <nonDerivativeTable><nonDerivativeTransaction>
    <securityTitle><value>Common Stock</value></securityTitle>
    <transactionDate><value>2026-07-14</value></transactionDate>
    <transactionCoding><transactionCode>P</transactionCode></transactionCoding>
    <transactionAmounts>
      <transactionShares><value>10</value></transactionShares>
      <transactionPricePerShare><value>2.5</value></transactionPricePerShare>
      <transactionAcquiredDisposedCode><value>A</value></transactionAcquiredDisposedCode>
    </transactionAmounts>
    <postTransactionAmounts><sharesOwnedFollowingTransaction><value>100</value></sharesOwnedFollowingTransaction></postTransactionAmounts>
    <ownershipNature><directOrIndirectOwnership><value>D</value></directOrIndirectOwnership></ownershipNature>
  </nonDerivativeTransaction></nonDerivativeTable>
</ownershipDocument>
</XML>
</SEC-DOCUMENT>"""
    hydrate_sec_filings(
        repo_root=tmp_path,
        registry=REGISTRY,
        index_path=index,
        forms=("4", "4/A"),
        limit=1,
        user_agent="Caerus Research test@example.com",
        fetcher=lambda *args, **kwargs: _response(body),
        sleeper=lambda _seconds: None,
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    master = tmp_path / "data/pit_universe/security_master.csv"
    master.parent.mkdir(parents=True)
    master.write_text(
        "security_id,cik,ticker,category,effective_start,effective_end\n"
        "SEC-123,123,TST,Domestic Common Stock,2010-01-01,\n",
        encoding="utf-8",
    )
    result = materialize_original_insider_events(
        tmp_path,
        generated_at=datetime(2026, 7, 16, 12, 0, tzinfo=UTC),
    )
    assert result["form4_original_quality_status"] == "PILOT_USABLE_ORIGINAL_FIRST"
    assert result["form4_original_event_rows"] == 1
    assert result["form4_original_eligible_purchase_rows"] == 1
    event_path = result["paths"]["events.jsonl.gz"]
    with gzip.open(event_path, "rt", encoding="utf-8") as stream:
        event = json.loads(next(stream))
    assert event["parse_status"] == "PASS_ORIGINAL_XML"
    assert event["security_id"] == "SEC-123"
    assert event["owner_cik"] == "0000000456"
    assert event["officer_title"] == "Chief Executive Officer"
    assert event["transaction_code"] == "P"
    assert event["acquired_disposed_code"] == "A"
    assert event["transaction_value"] == 25.0
    assert event["acceptance_datetime_utc"] == "2026-07-14T21:30:00+00:00"
    assert event["available_at"] == "2026-07-15T13:30:00+00:00"
    certification = json.loads(
        (tmp_path / "outputs/research/alpha_lab/provider_readiness/form4_event_tape_v1.json").read_text()
    )
    assert certification["status"] == "BLOCKED"
    assert "sample_only_not_full_history" in certification["blockers"]


def test_original_form4_stream_materialization_reads_each_tar_sequentially(tmp_path):
    index = (
        tmp_path
        / "outputs/research/alpha_lab/shared/form4_purchase_hydration_index.csv"
    )
    index.parent.mkdir(parents=True)
    index.write_text(
        "cik,company_name,form_type,filed_date,filename,index_year,index_quarter\n"
        "123,Issuer,4,2026-07-14,edgar/data/123/a.txt,2026,3\n",
        encoding="utf-8",
    )
    body = b"""<SEC-DOCUMENT>
<SEC-HEADER>ACCESSION NUMBER: 0000000123-26-000001
ACCEPTANCE-DATETIME: 20260714173000</SEC-HEADER>
<XML><ownershipDocument>
<documentType>4</documentType>
<issuer><issuerCik>123</issuerCik><issuerTradingSymbol>TST</issuerTradingSymbol></issuer>
<reportingOwner><reportingOwnerId><rptOwnerCik>456</rptOwnerCik><rptOwnerName>TEST PERSON</rptOwnerName></reportingOwnerId><reportingOwnerRelationship><isDirector>1</isDirector></reportingOwnerRelationship></reportingOwner>
<nonDerivativeTable><nonDerivativeTransaction><transactionDate><value>2026-07-14</value></transactionDate><transactionCoding><transactionCode>P</transactionCode></transactionCoding><transactionAmounts><transactionShares><value>10</value></transactionShares><transactionPricePerShare><value>2.5</value></transactionPricePerShare><transactionAcquiredDisposedCode><value>A</value></transactionAcquiredDisposedCode></transactionAmounts></nonDerivativeTransaction></nonDerivativeTable>
</ownershipDocument></XML></SEC-DOCUMENT>"""
    capture_sec_original_stream(
        repo_root=tmp_path,
        registry=REGISTRY,
        index_path=index,
        forms=("4", "4/A"),
        partition_size=1,
        user_agent="Caerus Research test@example.com",
        fetcher=lambda *args, **kwargs: _response(body),
        sleeper=lambda _seconds: None,
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    master = tmp_path / "data/pit_universe/security_master.csv"
    master.parent.mkdir(parents=True)
    master.write_text(
        "security_id,cik,ticker,category,effective_start,effective_end\n"
        "SEC-123,123,TST,Domestic Common Stock,2010-01-01,\n",
        encoding="utf-8",
    )
    result = materialize_original_insider_events(
        tmp_path,
        generated_at=datetime(2026, 7, 16, 12, 0, tzinfo=UTC),
    )
    assert result["form4_original_quality_status"] == "READY_FULL_HISTORY"
    assert result["form4_original_event_rows"] == 1


def test_original_form4_stream_materialization_deduplicates_discovery_rows(tmp_path):
    index = (
        tmp_path
        / "outputs/research/alpha_lab/shared/form4_purchase_hydration_index.csv"
    )
    index.parent.mkdir(parents=True)
    index.write_text(
        "cik,company_name,form_type,filed_date,filename,index_year,index_quarter\n"
        "123,Issuer,4,2026-07-14,edgar/data/123/a.txt,2026,3\n"
        "123,Issuer,4,2026-07-14,edgar/data/123/a.txt,2026,3\n",
        encoding="utf-8",
    )
    body = b"""<SEC-DOCUMENT>
<SEC-HEADER>ACCESSION NUMBER: 0000000123-26-000001
ACCEPTANCE-DATETIME: 20260714173000</SEC-HEADER>
<XML><ownershipDocument>
<documentType>4</documentType>
<issuer><issuerCik>123</issuerCik><issuerTradingSymbol>TST</issuerTradingSymbol></issuer>
<reportingOwner><reportingOwnerId><rptOwnerCik>456</rptOwnerCik><rptOwnerName>TEST PERSON</rptOwnerName></reportingOwnerId><reportingOwnerRelationship><isDirector>1</isDirector></reportingOwnerRelationship></reportingOwner>
<nonDerivativeTable><nonDerivativeTransaction><transactionDate><value>2026-07-14</value></transactionDate><transactionCoding><transactionCode>P</transactionCode></transactionCoding><transactionAmounts><transactionShares><value>10</value></transactionShares><transactionPricePerShare><value>2.5</value></transactionPricePerShare><transactionAcquiredDisposedCode><value>A</value></transactionAcquiredDisposedCode></transactionAmounts></nonDerivativeTransaction></nonDerivativeTable>
</ownershipDocument></XML></SEC-DOCUMENT>"""
    capture_sec_original_stream(
        repo_root=tmp_path,
        registry=REGISTRY,
        index_path=index,
        forms=("4", "4/A"),
        partition_size=1,
        user_agent="Caerus Research test@example.com",
        fetcher=lambda *args, **kwargs: _response(body),
        sleeper=lambda _seconds: None,
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    master = tmp_path / "data/pit_universe/security_master.csv"
    master.parent.mkdir(parents=True)
    master.write_text(
        "security_id,cik,ticker,category,effective_start,effective_end\n"
        "SEC-123,123,TST,Domestic Common Stock,2010-01-01,\n",
        encoding="utf-8",
    )
    result = materialize_original_insider_events(
        tmp_path,
        generated_at=datetime(2026, 7, 16, 12, 0, tzinfo=UTC),
    )
    quality = json.loads(result["paths"]["quality.json"].read_text(encoding="utf-8"))
    assert result["form4_original_event_rows"] == 1
    assert quality["counts"]["source_candidate_row_count"] == 2
    assert quality["counts"]["source_filing_count"] == 1
    assert quality["counts"]["duplicate_source_row_count"] == 1
    with gzip.open(result["paths"]["events.jsonl.gz"], "rt", encoding="utf-8") as stream:
        events = [json.loads(line) for line in stream]
    assert len({event["event_id"] for event in events}) == 1


def test_original_form4_stream_excludes_amended_issuer_and_certifies_clean_subset(
    tmp_path,
):
    index = (
        tmp_path
        / "outputs/research/alpha_lab/shared/form4_purchase_hydration_index.csv"
    )
    index.parent.mkdir(parents=True)
    index.write_text(
        "cik,company_name,form_type,filed_date,filename,index_year,index_quarter\n"
        "123,Issuer A,4,2026-07-14,edgar/data/123/a.txt,2026,3\n"
        "123,Issuer A,4/A,2026-07-15,edgar/data/123/b.txt,2026,3\n"
        "789,Issuer B,4,2026-07-14,edgar/data/789/c.txt,2026,3\n",
        encoding="utf-8",
    )

    def form4_body(
        *,
        accession: str,
        accepted: str,
        document_type: str,
        issuer_cik: str,
        ticker: str,
        owner_cik: str,
    ) -> bytes:
        return (
            "<SEC-DOCUMENT><SEC-HEADER>ACCESSION NUMBER: {accession}\n"
            "ACCEPTANCE-DATETIME: {accepted}</SEC-HEADER><XML>"
            "<ownershipDocument><documentType>{document_type}</documentType>"
            "<issuer><issuerCik>{issuer_cik}</issuerCik>"
            "<issuerTradingSymbol>{ticker}</issuerTradingSymbol></issuer>"
            "<reportingOwner><reportingOwnerId><rptOwnerCik>{owner_cik}</rptOwnerCik>"
            "<rptOwnerName>TEST PERSON</rptOwnerName></reportingOwnerId>"
            "<reportingOwnerRelationship><isDirector>1</isDirector>"
            "</reportingOwnerRelationship></reportingOwner>"
            "<nonDerivativeTable><nonDerivativeTransaction>"
            "<transactionDate><value>2026-07-14</value></transactionDate>"
            "<transactionCoding><transactionCode>P</transactionCode></transactionCoding>"
            "<transactionAmounts><transactionShares><value>10</value></transactionShares>"
            "<transactionPricePerShare><value>2.5</value></transactionPricePerShare>"
            "<transactionAcquiredDisposedCode><value>A</value>"
            "</transactionAcquiredDisposedCode></transactionAmounts>"
            "</nonDerivativeTransaction></nonDerivativeTable>"
            "</ownershipDocument></XML></SEC-DOCUMENT>"
        ).format(
            accession=accession,
            accepted=accepted,
            document_type=document_type,
            issuer_cik=issuer_cik,
            ticker=ticker,
            owner_cik=owner_cik,
        ).encode()

    bodies = {
        "a.txt": form4_body(
            accession="0000000123-26-000001",
            accepted="20260714173000",
            document_type="4",
            issuer_cik="123",
            ticker="AAA",
            owner_cik="456",
        ),
        "b.txt": form4_body(
            accession="0000000123-26-000002",
            accepted="20260715173000",
            document_type="4/A",
            issuer_cik="123",
            ticker="AAA",
            owner_cik="456",
        ),
        "c.txt": form4_body(
            accession="0000000789-26-000001",
            accepted="20260714173000",
            document_type="4",
            issuer_cik="789",
            ticker="BBB",
            owner_cik="654",
        ),
    }

    def fetcher(url, **_kwargs):
        return _response(next(body for name, body in bodies.items() if url.endswith(name)))

    capture_sec_original_stream(
        repo_root=tmp_path,
        registry=REGISTRY,
        index_path=index,
        forms=("4", "4/A"),
        partition_size=3,
        user_agent="Caerus Research test@example.com",
        fetcher=fetcher,
        sleeper=lambda _seconds: None,
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    master = tmp_path / "data/pit_universe/security_master.csv"
    master.parent.mkdir(parents=True)
    master.write_text(
        "security_id,cik,ticker,category,effective_start,effective_end\n"
        "SEC-123,123,AAA,Domestic Common Stock,2010-01-01,\n"
        "SEC-789,789,BBB,Domestic Common Stock,2010-01-01,\n",
        encoding="utf-8",
    )
    result = materialize_original_insider_events(
        tmp_path,
        generated_at=datetime(2026, 7, 16, 12, 0, tzinfo=UTC),
    )
    assert result["form4_original_quality_status"] == "READY_FULL_HISTORY"
    quality = json.loads(result["paths"]["quality.json"].read_text(encoding="utf-8"))
    assert quality["amendment_lineage_reconciled_by_exclusion"] is True
    assert quality["counts"]["amendment_issuer_count"] == 1
    assert quality["counts"]["amendment_excluded_event_row_count"] == 2
    with gzip.open(
        result["paths"]["events.jsonl.gz"], "rt", encoding="utf-8"
    ) as stream:
        events = [json.loads(line) for line in stream]
    assert [event["security_id"] for event in events] == ["SEC-789"]
    certification = json.loads(
        (
            tmp_path
            / "outputs/research/alpha_lab/provider_readiness/form4_event_tape_v1.json"
        ).read_text(encoding="utf-8")
    )
    assert certification["status"] == "READY"
    assert certification["historical_point_in_time_verified"] is True


def test_readiness_does_not_mislabel_8k_stream_as_form4_coverage(tmp_path, monkeypatch):
    monkeypatch.setenv("NASDAQ_DATA_LINK_API_KEY", "test-key")
    monkeypatch.setenv("SEC_USER_AGENT", "Caerus Research test@example.com")
    write_bundle(
        repo_root=tmp_path,
        source_id="sec_original_filings_stream",
        files={"partitions/part_00000.tar.gz": b"8k"},
        metadata={"forms": ["8-K", "8-K/A"], "candidate_count": 100},
        retrieved_at=datetime(2026, 7, 16, 12, 0, tzinfo=UTC),
    )
    readiness = build_readiness(
        repo_root=tmp_path,
        registry=REGISTRY,
        checked_at=datetime(2026, 7, 16, 13, 0, tzinfo=UTC),
    )["readiness"]
    assert readiness["sources"]["sec_original_8k_stream"]["status"] == "CAPTURED"
    assert readiness["sources"]["sec_original_form4_stream"]["status"] == "NOT_CAPTURED"


def _french_zip(header: str, row: str) -> bytes:
    stream = io.BytesIO()
    with zipfile.ZipFile(stream, "w") as archive:
        archive.writestr("data.csv", "description\n" + header + "\n" + row + "\n\nAnnual Factors\n")
    return stream.getvalue()


def _aqr_xlsx(value: float) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    for row_number in range(1, 19):
        sheet.cell(row=row_number, column=1, value="metadata")
    sheet.cell(row=19, column=1, value="DATE")
    sheet.cell(row=19, column=2, value="USA")
    sheet.cell(row=20, column=1, value="07/14/2026")
    sheet.cell(row=20, column=2, value=value)
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_french_parser_and_factor_bundle(tmp_path):
    ff5 = _french_zip(",Mkt-RF,SMB,HML,RMW,CMA,RF", "20260714,1,2,3,4,5,0.1")
    mom = _french_zip(",Mom", "20260714,6")
    reversal = _french_zip(",ST_Rev", "20260714,7")
    industry = _french_zip(",Industry", "20260714,8")
    aqr = _aqr_xlsx(0.01)

    def fetch(url, **kwargs):
        if "Momentum" in url:
            return _response(mom)
        if "ST_Reversal" in url:
            return _response(reversal)
        if "Industry" in url:
            return _response(industry)
        if "tuck.dartmouth" in url:
            return _response(ff5)
        return _response(aqr)

    header, rows = _parse_french_zip(ff5)
    assert "Mkt_RF" in header
    assert rows[0]["date"] == "20260714"
    result = collect_factors(
        repo_root=tmp_path,
        registry=REGISTRY,
        fetcher=fetch,
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    panel = result["paths"]["normalized/french_factor_panel.csv"].read_text()
    assert "MKT_RF" in panel
    assert "20260714" in panel
    combined = result["paths"]["normalized/factor_panel.csv"].read_text()
    assert "LOW_VOL_BAB" in combined
    assert "QMJ" in combined


def test_french_parser_skips_blank_comma_lines_and_trailing_cells():
    stream = io.BytesIO()
    with zipfile.ZipFile(stream, "w") as archive:
        archive.writestr(
            "momentum.csv",
            "description,,\n,,\n,Mom,\n20110103,1.25,\n\nAnnual Factors\n",
        )
    header, rows = _parse_french_zip(stream.getvalue())
    assert header == ["date", "Mom"]
    assert rows == [{"date": "20110103", "Mom": "1.25"}]


def test_fred_collection_uses_initial_release_output_without_key_leak(tmp_path, monkeypatch):
    monkeypatch.setenv("FRED_API_KEY", "secret-value")

    def fetch(url, **kwargs):
        assert "secret-value" in url
        payload = {
            "observations": [
                {
                    "date": "2026-07-01",
                    "value": "1.0",
                    "realtime_start": "2026-07-01",
                    "realtime_end": "2026-07-01",
                }
            ]
        }
        return _response(json.dumps(payload).encode())

    result = collect_fred_alfred(
        repo_root=tmp_path,
        registry=REGISTRY,
        fetcher=fetch,
        retrieved_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    assert result["manifest"]["metadata"]["fred_output_type"] == 4
    assert "secret-value" not in result["manifest_path"].read_text()


def test_eia_audit_accepts_free_bulk_and_enforces_size_gate(tmp_path):
    sizes = iter((4_000_000, 60_000_000, 290_000_000))
    result = audit_eia(
        repo_root=tmp_path,
        registry=REGISTRY,
        head_fetcher=lambda *args, **kwargs: _response(
            b"", {"content-length": str(next(sizes)), "last-modified": "today"}
        ),
        checked_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    rows = result["audit"]["datasets"]
    assert rows[0]["automatic_capture_allowed"] is True
    assert rows[-1]["automatic_capture_allowed"] is False
    assert result["audit"]["bulk_requires_key"] is False


def test_vendor_schema_gate_is_explicitly_not_pit_certification(tmp_path):
    sample = tmp_path / "sample.csv"
    sample.write_text("security_id,analyst_id\nA,1\n", encoding="utf-8")
    result = validate_vendor_sample(
        repo_root=tmp_path,
        kind="analyst_estimates",
        sample_path=sample,
        checked_at=datetime(2026, 7, 15, 12, 0, tzinfo=UTC),
    )
    assert result["gate"]["status"] == "BLOCKED_SCHEMA"
    assert result["gate"]["historical_point_in_time_verified"] is False


def test_data_spine_boundary_is_clean():
    attestation = build_boundary_attestation()
    assert attestation["production_boundary_status"] == "CLEAN"
    assert attestation["findings"] == []


def test_terminal_return_sensitivity_relabels_legacy_proxy_without_certifying_it(
    tmp_path,
):
    duckdb = pytest.importorskip("duckdb")
    panel = tmp_path / "outputs/research/pit_liquidity/pit_liquidity_panel.parquet"
    panel.parent.mkdir(parents=True)
    connection = duckdb.connect()
    try:
        connection.execute(
            """
            COPY (
              SELECT * FROM (VALUES
                ('SEC-1', DATE '2020-01-02', 10.0, NULL::DOUBLE),
                ('SEC-1', DATE '2020-01-03', 11.0, 0.10),
                ('SEC-2', DATE '2020-01-03', 20.0, 0.02)
              ) AS t(security_id, date, close, terminal_return)
            ) TO '{}' (FORMAT PARQUET)
            """.format(str(panel).replace("'", "''")),
        )
    finally:
        connection.close()
    master = tmp_path / "data/pit_universe/security_master.csv"
    master.parent.mkdir(parents=True)
    master.write_text(
        "security_id,effective_end,lastpricedate,ticker\n"
        "SEC-1,2020-01-03,2020-01-03,OLD\n"
        "SEC-2,,,LIVE\n",
        encoding="utf-8",
    )
    result = build_terminal_return_sensitivity(
        repo_root=tmp_path,
        generated_at=datetime(2026, 7, 23, 16, 0, tzinfo=UTC),
    )
    assert result["terminal_candidate_count"] == 1
    assert result["terminal_return_status"] == "SENSITIVITY_ONLY"
    assert result["quality"]["terminal_settlement_certified"] is False
    quality = json.loads(result["paths"]["quality.json"].read_text(encoding="utf-8"))
    assert quality["observed_return_lineage"].startswith("legacy_v2")
    connection = duckdb.connect()
    try:
        row = connection.execute(
            """
            SELECT provider_final_day_total_return, verified_terminal_return,
                   pessimistic_total_loss_return, zero_incremental_return,
                   use_in_primary_point_estimate
            FROM read_parquet(?)
            """,
            [str(result["paths"]["terminal_return_sensitivity.parquet"])],
        ).fetchone()
    finally:
        connection.close()
    assert row[0] == pytest.approx(0.10)
    assert row[1] is None
    assert row[2] == pytest.approx(-1.0)
    assert row[3] == pytest.approx(0.0)
    assert row[4] is False
    assert not (
        tmp_path
        / "outputs/research/alpha_lab/provider_readiness/pit_prices_liquidity_v1.json"
    ).exists()
    sensitivity_certification = json.loads(
        (
            tmp_path
            / "outputs/research/alpha_lab/provider_readiness/"
            "terminal_return_sensitivity_v1.json"
        ).read_text(encoding="utf-8")
    )
    assert sensitivity_certification["status"] == "READY"
    assert sensitivity_certification["historical_point_in_time_verified"] is True
    assert "cannot support an alpha claim alone" in sensitivity_certification[
        "methodology"
    ]


def test_market_materialization_does_not_mislabel_last_daily_return_as_settlement(
    tmp_path,
):
    duckdb = pytest.importorskip("duckdb")
    pytest.importorskip("pyarrow")
    sep_manifest = (
        tmp_path
        / "outputs/research/alpha_lab/data_spine/sharadar_sep_stream/test/manifest.json"
    )
    sep_path = sep_manifest.parent / "data/sep.csv.gz"
    sep_path.parent.mkdir(parents=True)
    with gzip.open(sep_path, "wt", encoding="utf-8") as stream:
        stream.write(
            "ticker,date,open,high,low,close,closeunadj,closeadj,volume,lastupdated\n"
            "OLD,2020-01-02,10,10,10,10,10,10,1000,2020-01-03\n"
            "OLD,2020-01-03,11,11,11,11,11,11,1000,2020-01-04\n"
        )
    sep_manifest.write_text(
        json.dumps(
            {
                "metadata": {
                    "table": "SHARADAR/SEP",
                    "pagination_complete": True,
                }
            }
        ),
        encoding="utf-8",
    )
    master = tmp_path / "data/pit_universe/security_master.csv"
    master.parent.mkdir(parents=True)
    master.write_text(
        "security_id,cik,sector,ticker,effective_start,effective_end\n"
        "SEC-1,1,Industrials,OLD,2010-01-01,2020-01-03\n",
        encoding="utf-8",
    )
    write_bundle(
        repo_root=tmp_path,
        source_id="sharadar_actions",
        files={
            "actions.jsonl": (
                b'{"ticker":"OLD","date":"2020-01-03","action":"dividend","value":0.1}\n'
            )
        },
        metadata={"test": True},
        retrieved_at=datetime(2026, 7, 23, 16, 0, tzinfo=UTC),
    )
    shared = tmp_path / "outputs/research/alpha_lab/shared"
    shared.mkdir(parents=True, exist_ok=True)
    (shared / "factor_panel.csv").write_text(
        "date,MKT_RF\n2020-01-02,0.01\n2020-01-03,0.01\n",
        encoding="utf-8",
    )
    with gzip.open(
        shared / "sec_companyfacts_compact.csv.gz", "wt", encoding="utf-8"
    ) as stream:
        stream.write(
            "cik,logical_fact,value,available_at,end,accession_number\n"
            "1,shares_outstanding,1000,2019-12-31T00:00:00Z,2019-12-31,A\n"
            "1,stockholders_equity,10000,2019-12-31T00:00:00Z,2019-12-31,A\n"
        )
    result = materialize_market_panels(
        repo_root=tmp_path,
        sep_manifest_path=sep_manifest,
    )
    connection = duckdb.connect()
    try:
        row = connection.execute(
            """
            SELECT closeadj, last_observed_total_return, delisting_return,
                   terminal_return
            FROM read_parquet(?)
            ORDER BY date DESC
            LIMIT 1
            """,
            [result["price_panel_path"]],
        ).fetchone()
    finally:
        connection.close()
    assert row[0] == pytest.approx(11.0)
    assert row[1] == pytest.approx(0.10)
    assert row[2] is None
    assert row[3] is None
    manifest = json.loads(
        (
            tmp_path / "outputs/research/pit_liquidity/manifest.json"
        ).read_text(encoding="utf-8")
    )
    assert manifest["schema_version"] == "caerus_pit_liquidity_panel_v3"
    assert manifest["verified_terminal_return_count"] == 0
    assert manifest["terminal_settlement_certified"] is False
    observed_certification = json.loads(
        (
            tmp_path
            / "outputs/research/alpha_lab/provider_readiness/"
            "pit_observed_prices_v1.json"
        ).read_text(encoding="utf-8")
    )
    assert observed_certification["status"] == "READY"
    assert observed_certification["historical_point_in_time_verified"] is True


def test_form4_amendment_policy_excludes_ambiguous_issuer_without_guessing(
    tmp_path,
):
    path = tmp_path / "events.jsonl.gz"
    rows = [
        {
            "event_id": "original-ambiguous",
            "issuer_cik": "0000000001",
            "amendment_lineage": "ORIGINAL",
            "eligible_open_market_purchase": True,
        },
        {
            "event_id": "amendment",
            "issuer_cik": "0000000001",
            "amendment_lineage": "AMENDMENT_UNRESOLVED",
            "eligible_open_market_purchase": True,
        },
        {
            "event_id": "clean-original",
            "issuer_cik": "0000000002",
            "amendment_lineage": "ORIGINAL",
            "eligible_open_market_purchase": True,
        },
    ]
    with gzip.open(path, "wt", encoding="utf-8") as stream:
        for row in rows:
            stream.write(json.dumps(row) + "\n")
    counts = _exclude_amendment_ambiguous_issuers(
        path,
        amendment_issuer_ciks={"0000000001"},
    )
    assert counts == {
        "retained_event_row_count": 1,
        "retained_eligible_purchase_row_count": 1,
        "excluded_event_row_count": 2,
        "excluded_eligible_purchase_row_count": 2,
    }
    with gzip.open(path, "rt", encoding="utf-8") as stream:
        retained = [json.loads(line) for line in stream]
    assert [row["event_id"] for row in retained] == ["clean-original"]
    assert (
        retained[0]["amendment_lineage"]
        == "ORIGINAL_ISSUER_AMENDMENT_FREE_IN_CAPTURE"
    )
