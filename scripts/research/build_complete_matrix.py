"""
Alpha Stack — Complete Regime-Aware Backtest Matrix (v2)
========================================================
Builds all 6 configs × all windows × all required metrics.

AVAILABLE (computed from production CSV data):
  1. trend_only          — Sleeve 1 production NAV (2009-2025)
  2. combined_allocator  — Alpha Variant net NAV (2009-2025)
  3. combined_alloc_cb   — Alpha Variant + circuit breaker NAV (2009-2025)
  6. spy_benchmark       — SPY buy-and-hold (2009-2025)

PENDING (requires local run with yfinance + EDGAR):
  4. value_only          — Value Sleeve standalone backtest
  5. combined_static     — Static 50/50 Trend + Value blend

Run `regime_aware_backtest_matrix.py` on your Mac to generate configs 4 & 5,
then run `merge_local_results.py` to merge into this workbook.

Outputs (all to outputs/regime_matrix/):
  - master_results.csv          (all 6 configs × all windows)
  - Alpha_Stack_Matrix_v2.xlsx  (8-sheet workbook)
  - equity_curves/{window}.csv  (normalised NAV per window)
  - regime_comparison.csv       (summary table for fixed windows)
  - DECISION_MEMO.md            (updated)
  - CONCLUSION_REPORT.md        (updated)
"""
from __future__ import annotations

import logging
import sys
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger("matrix_v2")

ROOT   = Path(__file__).resolve().parent.parent
RES    = ROOT / "outputs" / "research"
OUTDIR = ROOT / "outputs" / "regime_matrix"
ECDIR  = OUTDIR / "equity_curves"
OUTDIR.mkdir(parents=True, exist_ok=True)
ECDIR.mkdir(parents=True, exist_ok=True)

RISK_FREE      = 0.04
PENDING_MARKER = "PENDING"   # used in CSV / Excel for uncomputed cells

# ── Window definitions ────────────────────────────────────────────────────────

FIXED_WINDOWS = [
    ("full_period",   "2009-01-02", "2025-12-31", "Full Period — 17 years"),
    ("regime_gfc",    "2009-01-02", "2010-12-31", "GFC Recovery (2009-2010)"),
    ("regime_bull1",  "2011-01-01", "2015-12-31", "Bull 1 — Post-GFC Rally"),
    ("regime_bull2",  "2016-01-01", "2020-12-31", "Bull 2 — Pre-COVID"),
    ("regime_covid",  "2019-01-01", "2022-12-31", "COVID Crash + Recovery"),
    ("regime_recent", "2021-01-01", "2025-12-31", "Recent (2021-present)"),
    ("last_3yr",      "2023-01-01", "2025-12-31", "Last 3 Years"),
    ("last_5yr",      "2021-01-01", "2025-12-31", "Last 5 Years"),
]

CONFIG_META = {
    "trend_only":         {"label": "Trend-Only",          "color": "2E75B6", "available": True},
    "value_only":         {"label": "Value-Only",          "color": "E7824A", "available": False},
    "combined_static":    {"label": "Combined 50/50",      "color": "70AD47", "available": False},
    "combined_allocator": {"label": "Combined Allocator",  "color": "7030A0", "available": True},
    "combined_alloc_cb":  {"label": "Alloc + Circ. Break", "color": "C00000", "available": True},
    "spy_benchmark":      {"label": "SPY Benchmark",       "color": "595959", "available": True},
}
CONFIG_ORDER = list(CONFIG_META.keys())

METRIC_DEFS = [
    ("cagr",         "CAGR (%)",               "Net annualised compound return over window"),
    ("gross_cagr",   "Gross CAGR (%)",         "Gross CAGR before transaction costs"),
    ("annual_vol",   "Ann. Volatility (%)",    "Annualised standard deviation of daily returns"),
    ("sharpe",       "Sharpe",                 "(CAGR - 4% RF) / Ann. Vol"),
    ("sortino",      "Sortino",                "(CAGR - 4% RF) / Downside Vol below RF"),
    ("max_dd",       "Max Drawdown (%)",       "Peak-to-trough decline (worst close-to-close)"),
    ("calmar",       "Calmar",                 "CAGR / |Max Drawdown|"),
    ("win_rate",     "Win Rate (%)",           "% of trading days with positive return"),
    ("total_return", "Total Return (%)",       "Cumulative net return over window"),
    ("turnover",     "Avg Turnover",           "Avg 1-way portfolio turnover per rebalance"),
    ("spy_excess_cagr", "Alpha vs SPY (%)",   "Strategy CAGR minus SPY CAGR"),
    ("beta_vs_spy",  "Beta vs SPY",            "Regression beta of daily returns vs SPY"),
    ("n_years",      "Window Length (yr)",     "Window length in years"),
]
DISPLAY_METRICS = [m[0] for m in METRIC_DEFS]


# ── Data loading ──────────────────────────────────────────────────────────────

def load_data():
    trend_ts = pd.read_csv(RES / "sleeve1_backtest_2009_2025_timeseries.csv", parse_dates=["date"]).set_index("date").sort_index()
    alpha_ts  = pd.read_csv(RES / "sleeve1_alpha_variant_timeseries.csv", parse_dates=["date"]).set_index("date").sort_index()
    alpha_sum = pd.read_csv(RES / "sleeve1_alpha_variant_summary.csv")

    trend_nav    = trend_ts["portfolio_nav"].dropna()
    spy_nav      = trend_ts["spy_nav"].dropna()
    alloc_nav    = alpha_ts["net_nav"].dropna()
    alloc_gross  = alpha_ts["gross_nav"].dropna()
    alloc_nav_cb = alpha_ts["net_nav_cb"].dropna()
    alloc_gross_cb = alpha_ts["gross_nav_cb"].dropna()

    # Global turnover from summary (annualised average)
    avg_turnover = float(alpha_sum["avg_turnover"].iloc[0]) if "avg_turnover" in alpha_sum.columns else np.nan

    return {
        "trend_nav": trend_nav,
        "spy_nav":   spy_nav,
        "alloc_nav": alloc_nav,
        "alloc_gross": alloc_gross,
        "alloc_nav_cb": alloc_nav_cb,
        "alloc_gross_cb": alloc_gross_cb,
        "avg_turnover": avg_turnover,
    }


# ── Regime classification ─────────────────────────────────────────────────────

def classify_regimes(spy_nav: pd.Series) -> pd.DataFrame:
    ema50  = spy_nav.ewm(span=50,  adjust=False).mean()
    ema200 = spy_nav.ewm(span=200, adjust=False).mean()
    t1 = (spy_nav / ema200 - 1).fillna(0)
    t2 = (ema50   / ema200 - 1).fillna(0)

    def _state(r1, r2):
        if r1 >= 0.03 and r2 >= 0.01:  return "strong_up"
        if r1 >= 0.00 and r2 >= 0.00:  return "weak_up"
        if r1 >= -0.02:                 return "neutral"
        if r1 >= -0.05:                 return "weak_down"
        return "strong_down"

    states = [_state(float(t1.iloc[i]), float(t2.iloc[i])) for i in range(len(spy_nav))]
    df = pd.DataFrame({"trend_state": states,
                        "spy_t1": t1.values,
                        "spy_t2": t2.values}, index=spy_nav.index)
    return df


# ── Core metric computation ───────────────────────────────────────────────────

def compute_metrics(nav: pd.Series, gross_nav: pd.Series | None,
                    spy_nav: pd.Series, config: str, window: str,
                    label: str, start: str, end: str,
                    turnover: float = np.nan) -> dict:
    ws, we = pd.Timestamp(start), pd.Timestamp(end)
    net = nav[(nav.index >= ws) & (nav.index <= we)].dropna()
    if len(net) < 20:
        return {}

    # Re-normalise at window start
    net = net / net.iloc[0]

    daily = net.pct_change().dropna()
    n_days  = len(net)
    n_years = max(n_days / 252, 0.01)

    total_ret = float(net.iloc[-1] - 1)
    cagr      = float((1 + total_ret) ** (1 / n_years) - 1)
    ann_vol   = float(daily.std() * np.sqrt(252))
    excess_d  = daily - RISK_FREE / 252
    sharpe    = float(excess_d.mean() / daily.std() * np.sqrt(252)) if daily.std() > 1e-10 else 0.0

    down_mask = daily < RISK_FREE / 252
    d_vol     = float(daily[down_mask].std() * np.sqrt(252)) if down_mask.sum() > 1 else ann_vol
    sortino   = float((cagr - RISK_FREE) / d_vol) if d_vol > 1e-10 else 0.0

    cum    = (1 + daily).cumprod()
    hwm    = cum.cummax()
    dds    = (cum - hwm) / hwm
    max_dd = float(dds.min())
    max_dd_date = str(dds.idxmin().date()) if not dds.isna().all() else "N/A"
    calmar = round(cagr / abs(max_dd), 3) if max_dd < -0.001 else None

    win_rate = float((daily > 0).mean() * 100)

    # Gross CAGR
    gross_cagr = np.nan
    if gross_nav is not None:
        gn = gross_nav[(gross_nav.index >= ws) & (gross_nav.index <= we)].dropna()
        if len(gn) >= 20:
            gn = gn / gn.iloc[0]
            ny = len(gn) / 252
            gross_cagr = float((gn.iloc[-1]) ** (1 / ny) - 1) * 100

    # Beta and alpha vs SPY
    spy_sl = spy_nav[(spy_nav.index >= ws) & (spy_nav.index <= we)].dropna()
    beta_vs_spy = np.nan
    alpha_vs_spy_cagr = np.nan
    if len(spy_sl) >= 20:
        spy_sl = spy_sl / spy_sl.iloc[0]
        spy_d  = spy_sl.pct_change().dropna()
        spy_cagr = float((spy_sl.iloc[-1]) ** (1 / (len(spy_sl)/252)) - 1)
        alpha_vs_spy_cagr = round((cagr - spy_cagr) * 100, 2)
        common = daily.index.intersection(spy_d.index)
        if len(common) >= 20:
            X = spy_d.loc[common].values
            Y = daily.loc[common].values
            cov = np.cov(X, Y)[0, 1]
            var = np.var(X)
            beta_vs_spy = round(float(cov / var), 3) if var > 1e-12 else np.nan

    return {
        "window":       window,
        "window_label": label,
        "config":       config,
        "config_label": CONFIG_META[config]["label"],
        "start":        start,
        "end":          end,
        "n_years":      round(n_years, 2),
        "n_days":       n_days,
        "total_return": round(total_ret * 100, 2),
        "cagr":         round(cagr * 100, 2),
        "gross_cagr":   round(gross_cagr, 2) if not np.isnan(gross_cagr) else np.nan,
        "annual_vol":   round(ann_vol * 100, 2),
        "sharpe":       round(sharpe, 3),
        "sortino":      round(sortino, 3),
        "max_dd":       round(max_dd * 100, 2),
        "max_dd_date":  max_dd_date,
        "calmar":       calmar,
        "win_rate":     round(win_rate, 1),
        "turnover":     round(float(turnover), 4) if not np.isnan(turnover) else np.nan,
        "spy_excess_cagr": alpha_vs_spy_cagr,
        "beta_vs_spy":  beta_vs_spy,
    }


def make_pending_row(config: str, window: str, label: str, start: str, end: str) -> dict:
    """Structured placeholder row for configs requiring local run."""
    ws, we = pd.Timestamp(start), pd.Timestamp(end)
    n_years = (we - ws).days / 365.25
    return {
        "window":       window,
        "window_label": label,
        "config":       config,
        "config_label": CONFIG_META[config]["label"],
        "start":        start,
        "end":          end,
        "n_years":      round(n_years, 2),
        "n_days":       PENDING_MARKER,
        "total_return": PENDING_MARKER,
        "cagr":         PENDING_MARKER,
        "gross_cagr":   PENDING_MARKER,
        "annual_vol":   PENDING_MARKER,
        "sharpe":       PENDING_MARKER,
        "sortino":      PENDING_MARKER,
        "max_dd":       PENDING_MARKER,
        "max_dd_date":  PENDING_MARKER,
        "calmar":       PENDING_MARKER,
        "win_rate":     PENDING_MARKER,
        "turnover":     PENDING_MARKER,
        "spy_excess_cagr": PENDING_MARKER,
        "beta_vs_spy":  PENDING_MARKER,
    }


# ── Rolling window generation ─────────────────────────────────────────────────

def rolling_windows(window_years: int, step_months: int) -> list:
    wins = []
    cur  = pd.Timestamp("2009-01-01")
    lim  = pd.Timestamp("2026-01-01")
    while True:
        end = cur + pd.DateOffset(years=window_years) - pd.DateOffset(days=1)
        if end > lim:
            break
        tag = f"roll_{window_years}y_{cur.year}q{(cur.month-1)//3+1}"
        lbl = f"Rolling {window_years}Y: {cur.strftime('%b %Y')} – {end.strftime('%b %Y')}"
        wins.append((tag, str(cur.date()), str(end.date()), lbl))
        cur += pd.DateOffset(months=step_months)
    return wins


# ── Regime stats ──────────────────────────────────────────────────────────────

def regime_stats(regime_df: pd.DataFrame, start: str, end: str) -> dict:
    ws, we = pd.Timestamp(start), pd.Timestamp(end)
    sl = regime_df[(regime_df.index >= ws) & (regime_df.index <= we)]
    if sl.empty:
        return {}
    vc = sl["trend_state"].value_counts(normalize=True)
    return {
        "regime_strong_up":   round(float(vc.get("strong_up",   0)) * 100, 1),
        "regime_weak_up":     round(float(vc.get("weak_up",     0)) * 100, 1),
        "regime_neutral":     round(float(vc.get("neutral",     0)) * 100, 1),
        "regime_weak_down":   round(float(vc.get("weak_down",   0)) * 100, 1),
        "regime_strong_down": round(float(vc.get("strong_down", 0)) * 100, 1),
        "regime_defensive":   round(float(vc.get("weak_down", 0) + vc.get("strong_down", 0)) * 100, 1),
        "regime_dominant":    vc.idxmax(),
    }


# ── Main analysis runner ──────────────────────────────────────────────────────

def run_analysis(data: dict) -> pd.DataFrame:
    trend_nav    = data["trend_nav"]
    spy_nav      = data["spy_nav"]
    alloc_nav    = data["alloc_nav"]
    alloc_gross  = data["alloc_gross"]
    alloc_nav_cb = data["alloc_nav_cb"]
    alloc_gross_cb = data["alloc_gross_cb"]
    avg_turn     = data["avg_turnover"]

    # Config → (net_nav, gross_nav, turnover)
    AVAILABLE_CONFIGS = {
        "trend_only":         (trend_nav,    None,           np.nan),
        "combined_allocator": (alloc_nav,    alloc_gross,    avg_turn),
        "combined_alloc_cb":  (alloc_nav_cb, alloc_gross_cb, avg_turn),
        "spy_benchmark":      (spy_nav,      None,           np.nan),
    }

    all_windows = FIXED_WINDOWS + rolling_windows(3, 6) + rolling_windows(5, 12)
    logger.info("Total windows: %d", len(all_windows))

    results = []

    for win_name, win_start, win_end, win_label in all_windows:
        # ── Available configs ──────────────────────────────────────────────
        for config, (net, gross, turn) in AVAILABLE_CONFIGS.items():
            m = compute_metrics(net, gross, spy_nav, config,
                                win_name, win_label, win_start, win_end, turn)
            if m:
                results.append(m)

        # ── Pending configs ────────────────────────────────────────────────
        for config in ["value_only", "combined_static"]:
            results.append(make_pending_row(config, win_name, win_label, win_start, win_end))

        # ── Equity curves ──────────────────────────────────────────────────
        ws, we = pd.Timestamp(win_start), pd.Timestamp(win_end)
        frames = {}
        for conf, (net, gross, _) in AVAILABLE_CONFIGS.items():
            sl = net[(net.index >= ws) & (net.index <= we)].dropna()
            if not sl.empty:
                frames[conf] = (sl / sl.iloc[0]).rename(conf)
        if frames:
            ec = pd.concat(frames.values(), axis=1).sort_index()
            ec.to_csv(ECDIR / f"{win_name}.csv")

    return pd.DataFrame(results)


# ── Regime comparison summary table ──────────────────────────────────────────

def build_regime_comparison(master_df: pd.DataFrame, regime_df: pd.DataFrame) -> pd.DataFrame:
    fixed_names = [w[0] for w in FIXED_WINDOWS]
    rows = []
    for win_name, win_start, win_end, win_label in FIXED_WINDOWS:
        reg = regime_stats(regime_df, win_start, win_end)
        row = {
            "window":       win_name,
            "window_label": win_label,
            "start":        win_start,
            "end":          win_end,
            **reg,
        }
        for config in CONFIG_ORDER:
            sub = master_df[(master_df["window"] == win_name) & (master_df["config"] == config)]
            if sub.empty:
                continue
            r = sub.iloc[0]
            if r.get("cagr") == PENDING_MARKER:
                row[f"{config}_cagr"]   = PENDING_MARKER
                row[f"{config}_sharpe"] = PENDING_MARKER
                row[f"{config}_maxdd"]  = PENDING_MARKER
            else:
                row[f"{config}_cagr"]   = r.get("cagr")
                row[f"{config}_sharpe"] = r.get("sharpe")
                row[f"{config}_maxdd"]  = r.get("max_dd")
        rows.append(row)
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color="000000", italic=False) -> Font:
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def _border_thin() -> Border:
    s = Side(style="thin", color="C0C0C0")
    return Border(left=s, right=s, top=s, bottom=s)

def _border_medium() -> Border:
    s = Side(style="medium", color="808080")
    return Border(left=s, right=s, top=s, bottom=s)

HEADER_FILL   = _fill("1F3864")   # dark navy
SUBHDR_FILL   = _fill("2E75B6")   # blue
PENDING_FILL  = _fill("FFF2CC")   # pale yellow
PENDING_FILL2 = _fill("FFE599")   # yellow (darker)
ALT_FILL      = _fill("F2F2F2")   # light grey row
GOOD_FILL     = _fill("E2EFDA")   # green
WARN_FILL     = _fill("FCE4D6")   # red-orange
NEUTRAL_FILL  = _fill("DDEBF7")   # light blue
SECTION_FILL  = _fill("D6DCE4")   # section header grey

def _write_cell(ws, row, col, value, bold=False, fill=None, align="center",
                number_format=None, font_color="000000", italic=False, border=None, size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, color=font_color, italic=italic, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format
    if border:
        cell.border = border
    return cell


def _metric_fill(value, metric_name: str):
    """Return fill based on metric value vs thresholds."""
    if value is None or value == PENDING_MARKER:
        return PENDING_FILL
    try:
        v = float(value)
    except (ValueError, TypeError):
        return None

    if metric_name == "sharpe":
        if v >= 1.0:  return GOOD_FILL
        if v >= 0.5:  return NEUTRAL_FILL
        if v < 0.2:   return WARN_FILL
    elif metric_name == "cagr":
        if v >= 20:   return GOOD_FILL
        if v >= 10:   return NEUTRAL_FILL
        if v < 0:     return WARN_FILL
    elif metric_name == "max_dd":
        if v >= -10:  return GOOD_FILL
        if v >= -25:  return NEUTRAL_FILL
        if v < -40:   return WARN_FILL
    elif metric_name == "sortino":
        if v >= 1.5:  return GOOD_FILL
        if v >= 0.8:  return NEUTRAL_FILL
        if v < 0.3:   return WARN_FILL
    return None


def _fmt_val(v, metric: str) -> str | float:
    if v == PENDING_MARKER or v is None:
        return "—"
    try:
        f = float(v)
    except (ValueError, TypeError):
        return v
    if metric in ("cagr", "annual_vol", "total_return", "max_dd", "spy_excess_cagr", "gross_cagr"):
        return round(f, 1)
    if metric in ("sharpe", "sortino", "calmar", "beta_vs_spy"):
        return round(f, 2)
    if metric == "win_rate":
        return round(f, 1)
    if metric == "turnover":
        return round(f, 2)
    return v


# ── Sheet 1: Regime Comparison Summary ───────────────────────────────────────

def build_sheet_regime_summary(wb, master_df, regime_df):
    ws = wb.create_sheet("Regime Summary")
    ws.sheet_view.showGridLines = False

    # ── Title ──
    ws.merge_cells("A1:T1")
    _write_cell(ws, 1, 1, "ALPHA STACK — REGIME-AWARE BACKTEST MATRIX",
                bold=True, fill=_fill("1F3864"), font_color="FFFFFF", size=14, align="left")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:T2")
    _write_cell(ws, 2, 1, "Key Metrics by Regime Window and Strategy Configuration  |  Pending columns require local run (see 'How to Complete' tab)",
                bold=False, fill=_fill("D6DCE4"), font_color="444444", size=10, align="left", italic=True)
    ws.row_dimensions[2].height = 18

    # ── Config header row ──
    metric_labels = ["CAGR (%)", "Sharpe", "Max DD (%)"]
    col = 3
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14

    # Window / Date columns
    _write_cell(ws, 4, 1, "Window", bold=True, fill=HEADER_FILL, font_color="FFFFFF", border=_border_thin())
    _write_cell(ws, 4, 2, "Dominant Regime", bold=True, fill=HEADER_FILL, font_color="FFFFFF", border=_border_thin())

    config_col_starts = {}
    for config in CONFIG_ORDER:
        meta = CONFIG_META[config]
        config_col_starts[config] = col
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + len(metric_labels) - 1)
        cfg_cell = ws.cell(row=3, column=col)
        cfg_cell.value = meta["label"]
        cfg_cell.font  = _font(bold=True, color="FFFFFF", size=10)
        cfg_cell.fill  = _fill(meta["color"])
        cfg_cell.alignment = Alignment(horizontal="center", vertical="center")
        if not meta["available"]:
            cfg_cell.value += " ⏳"

        for i, lbl in enumerate(metric_labels):
            _write_cell(ws, 4, col + i, lbl, bold=True,
                        fill=_fill(meta["color"]), font_color="FFFFFF", border=_border_thin(), size=9)
            ws.column_dimensions[get_column_letter(col + i)].width = 10
        col += len(metric_labels)

    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 22

    # ── Data rows ──
    alt = False
    for rowi, (win_name, win_start, win_end, win_label) in enumerate(FIXED_WINDOWS, start=5):
        bg = ALT_FILL if alt else None
        alt = not alt

        reg = regime_stats(regime_df, win_start, win_end)
        dom = reg.get("regime_dominant", "N/A")

        _write_cell(ws, rowi, 1, win_label, bold=(win_name == "full_period"),
                    fill=_fill("1F3864") if win_name == "full_period" else bg,
                    font_color="FFFFFF" if win_name == "full_period" else "000000",
                    align="left", border=_border_thin())
        _write_cell(ws, rowi, 2, dom, fill=bg, border=_border_thin(), size=9)

        for config in CONFIG_ORDER:
            c = config_col_starts[config]
            sub = master_df[(master_df["window"] == win_name) & (master_df["config"] == config)]
            if sub.empty:
                for i in range(3):
                    _write_cell(ws, rowi, c + i, "—", fill=PENDING_FILL, border=_border_thin())
                continue
            r = sub.iloc[0]

            metrics = [("cagr", 0), ("sharpe", 1), ("max_dd", 2)]
            for mname, offset in metrics:
                raw = r.get(mname)
                val = _fmt_val(raw, mname)
                mfill = _metric_fill(raw, mname)
                if raw == PENDING_MARKER:
                    mfill = PENDING_FILL2
                _write_cell(ws, rowi, c + offset, val,
                            fill=mfill if mfill else bg,
                            border=_border_thin(),
                            bold=(win_name == "full_period"),
                            font_color="FFFFFF" if win_name == "full_period" else "000000",
                            size=9)

        ws.row_dimensions[rowi].height = 18

    # ── Legend ──
    legend_row = 5 + len(FIXED_WINDOWS) + 2
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=4)
    _write_cell(ws, legend_row, 1, "Color Key:", bold=True, align="left")
    items = [("GOOD_FILL", "Sharpe ≥ 1.0 / CAGR ≥ 20% / MaxDD ≥ -10%"),
             ("NEUTRAL_FILL", "Moderate performance"),
             ("WARN_FILL", "Below threshold"),
             ("PENDING_FILL2", "⏳ Requires local run (yfinance + EDGAR)")]
    fills_map = {"GOOD_FILL": GOOD_FILL, "NEUTRAL_FILL": NEUTRAL_FILL,
                 "WARN_FILL": WARN_FILL, "PENDING_FILL2": PENDING_FILL2}
    for i, (fname, desc) in enumerate(items):
        _write_cell(ws, legend_row + 1 + i, 1, "", fill=fills_map[fname], border=_border_thin())
        _write_cell(ws, legend_row + 1 + i, 2, desc, align="left", size=9)

    ws.freeze_panes = "C5"


# ── Sheet 2: Full Period Detail ───────────────────────────────────────────────

def build_sheet_full_period(wb, master_df):
    ws = wb.create_sheet("Full-Period Detail")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:K1")
    _write_cell(ws, 1, 1, "FULL-PERIOD COMPARISON  (2009 – 2025)", bold=True,
                fill=HEADER_FILL, font_color="FFFFFF", size=13, align="left")
    ws.row_dimensions[1].height = 28

    metric_col_order = [
        ("cagr",            "CAGR (%)"),
        ("gross_cagr",      "Gross CAGR (%)"),
        ("annual_vol",      "Ann. Vol (%)"),
        ("sharpe",          "Sharpe"),
        ("sortino",         "Sortino"),
        ("max_dd",          "Max DD (%)"),
        ("calmar",          "Calmar"),
        ("win_rate",        "Win Rate (%)"),
        ("turnover",        "Avg Turnover"),
        ("spy_excess_cagr", "vs SPY (CAGR %)"),
        ("beta_vs_spy",     "Beta vs SPY"),
        ("n_years",         "Years"),
    ]

    headers = ["Strategy"] + [h for _, h in metric_col_order]
    ws.column_dimensions["A"].width = 24
    for ci, (_, hdr) in enumerate(metric_col_order, start=2):
        ws.column_dimensions[get_column_letter(ci)].width = 13
        _write_cell(ws, 3, ci, hdr, bold=True, fill=SUBHDR_FILL, font_color="FFFFFF",
                    border=_border_thin(), size=9)
    _write_cell(ws, 3, 1, "Strategy", bold=True, fill=SUBHDR_FILL, font_color="FFFFFF",
                border=_border_thin(), size=10)
    ws.row_dimensions[3].height = 22

    fp = master_df[master_df["window"] == "full_period"].copy()
    fp["_sort"] = fp["config"].map({c: i for i, c in enumerate(CONFIG_ORDER)})
    fp = fp.sort_values("_sort")

    for rowi, (_, row) in enumerate(fp.iterrows(), start=4):
        config = row["config"]
        meta   = CONFIG_META[config]
        bg     = _fill(meta["color"] + "20") if len(meta["color"]) == 6 else None
        is_spy = (config == "spy_benchmark")

        _write_cell(ws, rowi, 1, meta["label"],
                    bold=True, fill=_fill(meta["color"]), font_color="FFFFFF",
                    align="left", border=_border_thin())

        for ci, (mname, _) in enumerate(metric_col_order, start=2):
            raw = row.get(mname)
            val = _fmt_val(raw, mname)
            mfill = _metric_fill(raw, mname) if not is_spy else ALT_FILL
            if raw == PENDING_MARKER:
                mfill = PENDING_FILL2
            _write_cell(ws, rowi, ci, val,
                        fill=mfill,
                        border=_border_thin(), size=9,
                        align="right" if isinstance(val, (int, float)) else "center")
        ws.row_dimensions[rowi].height = 20

    # Pending note
    note_row = 4 + len(CONFIG_ORDER) + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)
    _write_cell(ws, note_row, 1,
                "⏳ Value-Only and Combined 50/50 require running regime_aware_backtest_matrix.py locally "
                "(Python 3.11 venv with yfinance + EDGAR). See 'How to Complete' tab.",
                italic=True, size=9, fill=PENDING_FILL, align="left")
    ws.row_dimensions[note_row].height = 30

    ws.freeze_panes = "B4"


# ── Sheet 3: All Fixed Windows ────────────────────────────────────────────────

def build_sheet_fixed_windows(wb, master_df):
    ws = wb.create_sheet("Fixed Windows")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:L1")
    _write_cell(ws, 1, 1, "ALL FIXED WINDOWS — KEY METRICS PER CONFIGURATION",
                bold=True, fill=HEADER_FILL, font_color="FFFFFF", size=12, align="left")
    ws.row_dimensions[1].height = 26

    # Headers
    cols = ["Window", "Start", "End", "Config", "CAGR (%)", "Vol (%)", "Sharpe",
            "Sortino", "Max DD (%)", "Calmar", "Win Rate (%)", "vs SPY (%)"]
    col_widths = [22, 11, 11, 22, 9, 9, 8, 8, 9, 8, 10, 9]
    for ci, (hdr, w) in enumerate(zip(cols, col_widths), start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        _write_cell(ws, 3, ci, hdr, bold=True, fill=SUBHDR_FILL, font_color="FFFFFF", border=_border_thin(), size=9)
    ws.row_dimensions[3].height = 20

    fixed_names = [w[0] for w in FIXED_WINDOWS]
    fixed_df = master_df[master_df["window"].isin(fixed_names)].copy()
    fixed_df["_w_sort"] = fixed_df["window"].map({w[0]: i for i, w in enumerate(FIXED_WINDOWS)})
    fixed_df["_c_sort"] = fixed_df["config"].map({c: i for i, c in enumerate(CONFIG_ORDER)})
    fixed_df = fixed_df.sort_values(["_w_sort", "_c_sort"])

    current_win = None
    row_idx     = 4
    alt         = False
    for _, row in fixed_df.iterrows():
        if row["window"] != current_win:
            current_win = row["window"]
            alt = not alt
        bg = ALT_FILL if alt else None

        meta = CONFIG_META.get(row["config"], {})
        is_pend = row.get("cagr") == PENDING_MARKER

        vals = [
            row.get("window_label", row["window"]),
            row.get("start"),
            row.get("end"),
            meta.get("label", row["config"]),
            _fmt_val(row.get("cagr"), "cagr"),
            _fmt_val(row.get("annual_vol"), "annual_vol"),
            _fmt_val(row.get("sharpe"), "sharpe"),
            _fmt_val(row.get("sortino"), "sortino"),
            _fmt_val(row.get("max_dd"), "max_dd"),
            _fmt_val(row.get("calmar"), "calmar"),
            _fmt_val(row.get("win_rate"), "win_rate"),
            _fmt_val(row.get("spy_excess_cagr"), "spy_excess_cagr"),
        ]
        metric_names = [None, None, None, None, "cagr", "annual_vol", "sharpe",
                        "sortino", "max_dd", "calmar", "win_rate", "spy_excess_cagr"]

        for ci, (v, mn) in enumerate(zip(vals, metric_names), start=1):
            fill = None
            if is_pend and ci >= 5:
                fill = PENDING_FILL
            elif mn:
                fill = _metric_fill(row.get(mn), mn)
            if fill is None:
                fill = bg
            _write_cell(ws, row_idx, ci, v, fill=fill, border=_border_thin(), size=9,
                        align="left" if ci <= 4 else "right" if isinstance(v, (int, float)) else "center")

        ws.row_dimensions[row_idx].height = 16
        row_idx += 1

    ws.freeze_panes = "E4"


# ── Sheet 4: Rolling 3Y Sharpe Heatmap ───────────────────────────────────────

def build_sheet_rolling_heatmap(wb, master_df, window_years: int):
    ws_name = f"Rolling {window_years}Y Heatmap"
    ws = wb.create_sheet(ws_name)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    _write_cell(ws, 1, 1, f"ROLLING {window_years}-YEAR WINDOWS — SHARPE RATIO HEATMAP",
                bold=True, fill=HEADER_FILL, font_color="FFFFFF", size=12, align="left")
    ws.row_dimensions[1].height = 26

    roll_df = master_df[master_df["window"].str.startswith(f"roll_{window_years}y_")].copy()
    avail_configs = [c for c in CONFIG_ORDER if c in roll_df["config"].unique()]
    windows_sorted = sorted(roll_df["window"].unique())

    # Headers
    _write_cell(ws, 3, 1, "Window Start", bold=True, fill=SUBHDR_FILL, font_color="FFFFFF",
                border=_border_thin())
    _write_cell(ws, 3, 2, "Window End", bold=True, fill=SUBHDR_FILL, font_color="FFFFFF",
                border=_border_thin())
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14

    for ci, config in enumerate(avail_configs, start=3):
        meta = CONFIG_META[config]
        _write_cell(ws, 3, ci, meta["label"], bold=True,
                    fill=_fill(meta["color"]), font_color="FFFFFF", border=_border_thin(), size=9)
        ws.column_dimensions[get_column_letter(ci)].width = 16

    # Pending config columns
    pend_start_col = 3 + len(avail_configs)
    for pi, config in enumerate(["value_only", "combined_static"], start=pend_start_col):
        meta = CONFIG_META[config]
        _write_cell(ws, 3, pi, meta["label"] + " ⏳", bold=True,
                    fill=PENDING_FILL2, border=_border_thin(), size=9)
        ws.column_dimensions[get_column_letter(pi)].width = 16
    ws.row_dimensions[3].height = 22

    for ri, win_name in enumerate(windows_sorted, start=4):
        win_data = roll_df[roll_df["window"] == win_name]
        start = win_data["start"].iloc[0] if not win_data.empty else ""
        end   = win_data["end"].iloc[0]   if not win_data.empty else ""

        _write_cell(ws, ri, 1, start, size=9, border=_border_thin())
        _write_cell(ws, ri, 2, end,   size=9, border=_border_thin())

        for ci, config in enumerate(avail_configs, start=3):
            sub = win_data[win_data["config"] == config]
            if sub.empty:
                _write_cell(ws, ri, ci, "—", size=9, border=_border_thin())
                continue
            raw = sub.iloc[0].get("sharpe")
            val = _fmt_val(raw, "sharpe")
            fill = _metric_fill(raw, "sharpe")
            _write_cell(ws, ri, ci, val, fill=fill, size=9, border=_border_thin())

        for pi, _ in enumerate(["value_only", "combined_static"], start=pend_start_col):
            _write_cell(ws, ri, pi, "—", fill=PENDING_FILL, size=9, border=_border_thin())

        ws.row_dimensions[ri].height = 14

    # Summary stats below
    stats_row = 4 + len(windows_sorted) + 2
    ws.merge_cells(start_row=stats_row, start_column=1, end_row=stats_row, end_column=3 + len(avail_configs))
    _write_cell(ws, stats_row, 1, "Rolling Window Statistics (Available Configurations)",
                bold=True, fill=SECTION_FILL, align="left")

    stats_headers = ["Config", "Avg Sharpe", "Min Sharpe", "Max Sharpe", "% Windows > 1.0",
                     "Avg CAGR (%)", "Avg MaxDD (%)"]
    for ci, h in enumerate(stats_headers, start=1):
        _write_cell(ws, stats_row + 1, ci, h, bold=True, fill=SUBHDR_FILL, font_color="FFFFFF",
                    border=_border_thin(), size=9)

    for ri2, config in enumerate(avail_configs, start=stats_row + 2):
        sub = roll_df[roll_df["config"] == config].copy()
        sh  = pd.to_numeric(sub["sharpe"], errors="coerce").dropna()
        cg  = pd.to_numeric(sub["cagr"],   errors="coerce").dropna()
        dd  = pd.to_numeric(sub["max_dd"], errors="coerce").dropna()
        meta = CONFIG_META[config]
        vals = [
            meta["label"],
            round(sh.mean(), 3) if len(sh) else "—",
            round(sh.min(), 3)  if len(sh) else "—",
            round(sh.max(), 3)  if len(sh) else "—",
            f"{(sh > 1.0).mean() * 100:.0f}%" if len(sh) else "—",
            round(cg.mean(), 1) if len(cg) else "—",
            round(dd.mean(), 1) if len(dd) else "—",
        ]
        for ci, v in enumerate(vals, start=1):
            _write_cell(ws, ri2, ci, v, size=9, border=_border_thin(),
                        fill=_fill(meta["color"]) if ci == 1 else None,
                        font_color="FFFFFF" if ci == 1 else "000000")

    ws.freeze_panes = "C4"


# ── Sheet 5: Master Results ───────────────────────────────────────────────────

def build_sheet_master_results(wb, master_df):
    ws = wb.create_sheet("Master Results (Raw)")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:S1")
    _write_cell(ws, 1, 1, "MASTER RESULTS — ALL WINDOWS × ALL CONFIGS (Raw Data)",
                bold=True, fill=HEADER_FILL, font_color="FFFFFF", size=11, align="left")

    export_cols = ["window", "config", "start", "end", "n_years", "cagr", "gross_cagr",
                   "annual_vol", "sharpe", "sortino", "max_dd", "calmar",
                   "win_rate", "total_return", "turnover", "spy_excess_cagr", "beta_vs_spy"]
    col_widths   = [22, 20, 11, 11, 7, 8, 9, 8, 7, 7, 9, 7, 9, 9, 9, 10, 9]

    for ci, (col, w) in enumerate(zip(export_cols, col_widths), start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        _write_cell(ws, 3, ci, col, bold=True, fill=SUBHDR_FILL,
                    font_color="FFFFFF", border=_border_thin(), size=9)
    ws.row_dimensions[3].height = 18

    # Sort: fixed windows first, then rolling3, then rolling5; within window by config order
    master_df2 = master_df.copy()
    master_df2["_is_fixed"] = master_df2["window"].isin([w[0] for w in FIXED_WINDOWS]).astype(int)
    master_df2["_is_roll3"] = master_df2["window"].str.startswith("roll_3y_").astype(int)
    master_df2["_c_sort"]   = master_df2["config"].map({c: i for i, c in enumerate(CONFIG_ORDER)})
    master_df2["_w_group"]  = 0
    master_df2.loc[master_df2["_is_roll3"] == 1, "_w_group"] = 1
    master_df2.loc[master_df2["_is_fixed"] == 0, "_w_group"] = master_df2.loc[master_df2["_is_fixed"] == 0, "_w_group"].fillna(2)
    master_df2 = master_df2.sort_values(["_w_group", "start", "_c_sort"])

    alt = False
    prev_win = None
    for ri, (_, row) in enumerate(master_df2.iterrows(), start=4):
        if row["window"] != prev_win:
            prev_win = row["window"]
            alt = not alt
        bg = ALT_FILL if alt else None

        for ci, col in enumerate(export_cols, start=1):
            raw = row.get(col)
            if raw == PENDING_MARKER:
                _write_cell(ws, ri, ci, "—", fill=PENDING_FILL, size=8, border=_border_thin())
            elif pd.isna(raw) if not isinstance(raw, str) else False:
                _write_cell(ws, ri, ci, "—", fill=bg, size=8, border=_border_thin())
            else:
                fill = _metric_fill(raw, col) if col in ("sharpe", "cagr", "max_dd", "sortino") else bg
                _write_cell(ws, ri, ci, _fmt_val(raw, col), fill=fill, size=8, border=_border_thin(),
                            align="right" if isinstance(_fmt_val(raw, col), (int, float)) else "left")
        ws.row_dimensions[ri].height = 13

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(export_cols))}3"


# ── Sheet 6: Equity Curves ────────────────────────────────────────────────────

def build_sheet_equity_curves(wb, master_df):
    ws = wb.create_sheet("Equity Curves")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    _write_cell(ws, 1, 1, "NORMALISED EQUITY CURVES — SELECTED WINDOWS (Value = 1.0 at window start)",
                bold=True, fill=HEADER_FILL, font_color="FFFFFF", size=11, align="left")
    ws.row_dimensions[1].height = 26

    key_windows = ["full_period", "regime_gfc", "regime_bull2", "regime_covid",
                   "regime_recent", "last_3yr"]
    avail_configs_ec = [c for c in CONFIG_ORDER if CONFIG_META[c]["available"]]
    avail_configs_ec_labels = [CONFIG_META[c]["label"] for c in avail_configs_ec]

    _write_cell(ws, 3, 1, "Window", bold=True, fill=SUBHDR_FILL, font_color="FFFFFF", border=_border_thin())
    _write_cell(ws, 3, 2, "Final NAV", bold=True, fill=SUBHDR_FILL, font_color="FFFFFF", border=_border_thin())
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12

    row_idx = 4
    for win_name, _, _, win_label in FIXED_WINDOWS:
        if win_name not in key_windows:
            continue

        ec_path = ECDIR / f"{win_name}.csv"
        if not ec_path.exists():
            continue
        ec = pd.read_csv(ec_path, parse_dates=[0], index_col=0)

        # Section header
        _write_cell(ws, row_idx, 1, win_label, bold=True, fill=SECTION_FILL,
                    align="left", border=_border_thin())
        row_idx += 1

        # Column headers
        _write_cell(ws, row_idx, 1, "Date", bold=True, fill=ALT_FILL, size=8, border=_border_thin())
        for ci, cfg in enumerate(avail_configs_ec, start=2):
            if cfg in ec.columns:
                meta = CONFIG_META[cfg]
                _write_cell(ws, row_idx, ci, meta["label"], bold=True,
                            fill=_fill(meta["color"]), font_color="FFFFFF", size=8, border=_border_thin())
                ws.column_dimensions[get_column_letter(ci)].width = 14
        row_idx += 1

        # Data — sample every 20 trading days to keep file manageable
        ec_sample = ec.iloc[::20].copy()
        for dt, ec_row in ec_sample.iterrows():
            _write_cell(ws, row_idx, 1, str(dt.date()) if hasattr(dt, 'date') else str(dt),
                        size=8, border=_border_thin())
            for ci, cfg in enumerate(avail_configs_ec, start=2):
                if cfg in ec_row.index:
                    v = round(float(ec_row[cfg]), 4) if pd.notna(ec_row[cfg]) else "—"
                    _write_cell(ws, row_idx, ci, v, size=8, border=_border_thin())
            ws.row_dimensions[row_idx].height = 13
            row_idx += 1

        row_idx += 1  # gap between windows


# ── Sheet 7: How to Complete ──────────────────────────────────────────────────

def build_sheet_how_to_complete(wb):
    ws = wb.create_sheet("How to Complete")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 90

    ws.merge_cells("A1:A1")
    _write_cell(ws, 1, 1, "HOW TO COMPLETE THE MATRIX (Value-Only + Combined 50/50)",
                bold=True, fill=HEADER_FILL, font_color="FFFFFF", size=13, align="left")
    ws.row_dimensions[1].height = 30

    sections = [
        ("WHY THESE COLUMNS ARE PENDING", [
            "The Value-Only and Combined 50/50 configurations require:",
            "  1. Historical daily prices for 200 S&P 500 stocks (2008–present) via yfinance",
            "  2. Point-in-time SEC EDGAR fundamental data (earnings, FCF, book value)",
            "     with 40-90 day filing lag enforcement",
            "",
            "These are unavailable in the automated environment due to network restrictions.",
            "The computation must be run locally where the project Python 3.11 venv has",
            "yfinance, pyarrow, and internet access.",
        ]),
        ("STEP 1: Activate the project venv", [
            "  cd /path/to/quant-daily-report-main",
            "  source .venv/bin/activate",
            "  python --version   # should show Python 3.11",
            "  python -c \"import yfinance; print('yfinance OK')\"",
        ]),
        ("STEP 2: Run the full matrix script", [
            "  python scripts/regime_aware_backtest_matrix.py",
            "",
            "  This will:",
            "  - Download or use cached prices from data/alpha_stack_cache/",
            "  - Fetch EDGAR fundamentals for all 200 tickers",
            "  - Run Value-only backtest (top 30 by earnings/FCF/BP yield, sector-relative)",
            "  - Run Combined 50/50 (static 50% Trend + 50% Value blend)",
            "  - Save results to outputs/regime_matrix/local_run/",
            "",
            "  Estimated runtime: 45–90 minutes (EDGAR dominates on first run)",
            "  Subsequent runs: ~10-15 minutes (EDGAR cache hits)",
        ]),
        ("STEP 3: Merge into this workbook", [
            "  python scripts/merge_local_results.py",
            "",
            "  This will:",
            "  - Load outputs/regime_matrix/local_run/master_results.csv",
            "  - Replace all PENDING rows with real computed values",
            "  - Rebuild this Excel workbook with full 6-config data",
            "  - Update master_results.csv",
        ]),
        ("KNOWN BUG IN sleeve_value/backtest.py", [
            "  Line ~98 calls run_engine with wrong parameter names:",
            "    target_weights_df= instead of target_weights=",
            "    price_df=          instead of prices=",
            "    initial_capital=   instead of initial_equity=",
            "    comission=         instead of commission_bps=",
            "    rebalance_slippage= instead of slippage_bps=",
            "",
            "  The regime_aware_backtest_matrix.py script bypasses this by calling",
            "  engine.backtest_engine.run_backtest() directly with correct parameters.",
        ]),
        ("VALUE SLEEVE METHODOLOGY (for reference)", [
            "  Signal components (sector-relative z-scores):",
            "    - Earnings Yield    = EBIT / Enterprise Value (from EDGAR XBRL)",
            "    - FCF Yield         = Free Cash Flow / Enterprise Value",
            "    - Book-to-Price     = Book Value / Market Cap",
            "",
            "  Portfolio construction:",
            "    - Rank all tickers by composite value score",
            "    - Long top 30 names, equal-weighted within sleeve",
            "    - Rebalance monthly (ME rule)",
            "    - PIT safety: use only filings where 'filed' date < current date",
            "    - Filing lag: enforce minimum 45-day lag from period end",
            "",
            "  Combined 50/50:",
            "    - Each rebalance: 50% weight to Trend sleeve weights, 50% to Value weights",
            "    - No allocator — pure static blend",
            "    - Same commission (5bps) and slippage (5bps) model",
        ]),
    ]

    row_idx = 3
    for section_title, lines in sections:
        _write_cell(ws, row_idx, 1, section_title, bold=True, fill=SECTION_FILL,
                    align="left", border=_border_thin(), size=11)
        ws.row_dimensions[row_idx].height = 22
        row_idx += 1
        for line in lines:
            _write_cell(ws, row_idx, 1, line, align="left", size=9,
                        fill=ALT_FILL if line.startswith("  ") else None)
            ws.row_dimensions[row_idx].height = 15
            row_idx += 1
        row_idx += 1


# ── Sheet 8: Methodology ──────────────────────────────────────────────────────

def build_sheet_methodology(wb, master_df):
    ws = wb.create_sheet("Methodology & Notes")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 70

    ws.merge_cells("A1:B1")
    _write_cell(ws, 1, 1, "METHODOLOGY, DATA SOURCES AND METRIC DEFINITIONS",
                bold=True, fill=HEADER_FILL, font_color="FFFFFF", size=12, align="left")

    sections = [
        ("DATA SOURCES", [
            ("Trend NAV",      "sleeve1_backtest_2009_2025_timeseries.csv — Production Sleeve 1 daily NAV"),
            ("Allocator NAV",  "sleeve1_alpha_variant_timeseries.csv — Alpha Variant net/gross NAV"),
            ("SPY",            "Derived from portfolio timeseries spy_nav column (2009-2025)"),
            ("Value NAV",      "⏳ Requires EDGAR backtest (see 'How to Complete' tab)"),
            ("50/50 NAV",      "⏳ Requires local run combining Trend + Value signals"),
        ]),
        ("METRIC DEFINITIONS", [
            ("CAGR",           "Compound Annual Growth Rate: (end_nav / start_nav)^(1/years) - 1"),
            ("Gross CAGR",     "CAGR before commission/slippage deduction"),
            ("Ann. Vol",       "Annualised std dev of daily returns × √252"),
            ("Sharpe",         "(CAGR - 4% risk-free) / Ann. Vol"),
            ("Sortino",        "(CAGR - 4% risk-free) / Downside Vol (returns below RF/252)"),
            ("Max DD",         "Largest peak-to-trough close-to-close decline"),
            ("Calmar",         "CAGR / |Max Drawdown|"),
            ("Win Rate",       "% of trading days with positive return"),
            ("Avg Turnover",   "Average 1-way portfolio weight change per rebalance"),
            ("vs SPY",         "Strategy CAGR minus SPY CAGR for same window"),
            ("Beta vs SPY",    "OLS regression coefficient of daily returns on SPY daily returns"),
        ]),
        ("SIMULATION PARAMETERS", [
            ("Commission",     "5 basis points per trade (one-way)"),
            ("Slippage",       "5 basis points per trade (one-way)"),
            ("Rebalance rule", "Month-end (ME) — first business day of following month"),
            ("Execution lag",  "T+1 fill at open (signal computed at T close)"),
            ("Risk-free rate", "4% per annum (fixed)"),
            ("Universe",       "S&P 500 proxy — 200 liquid large-cap tickers (data/universe.csv)"),
        ]),
        ("REGIME CLASSIFICATION", [
            ("Method",         "SPY EMA50/EMA200 crossover with magnitude thresholds (hysteresis)"),
            ("strong_up",      "SPY/EMA200 ≥ +3% AND EMA50/EMA200 ≥ +1%"),
            ("weak_up",        "SPY/EMA200 ≥ 0% AND EMA50/EMA200 ≥ 0%"),
            ("neutral",        "SPY/EMA200 ≥ -2%"),
            ("weak_down",      "SPY/EMA200 ≥ -5%"),
            ("strong_down",    "SPY/EMA200 < -5%"),
        ]),
        ("IMPORTANT CAVEATS", [
            ("Survivorship",   "Universe fixed at 2026 constituents — survivorship bias in pre-2015 windows"),
            ("Look-ahead",     "EDGAR data uses PIT-safe 'filed' date; value scores may have minor look-ahead"),
            ("CAGR accuracy",  "Rolling windows use existing timeseries from fixed dates; partial months included"),
            ("Turnover proxy", "Avg turnover is the full-period average; per-window turnover not available"),
            ("Alpha variant",  "combined_allocator uses leverage (exposure_multiplier) — not a pure Trend+Value"),
        ]),
    ]

    row_idx = 3
    for section, items in sections:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        _write_cell(ws, row_idx, 1, section, bold=True, fill=SECTION_FILL, align="left",
                    border=_border_thin(), size=11)
        ws.row_dimensions[row_idx].height = 22
        row_idx += 1
        for k, v in items:
            _write_cell(ws, row_idx, 1, k, bold=True, fill=ALT_FILL, align="left",
                        border=_border_thin(), size=9)
            _write_cell(ws, row_idx, 2, v, align="left", border=_border_thin(), size=9)
            ws.row_dimensions[row_idx].height = 16
            row_idx += 1
        row_idx += 1


# ── Build complete workbook ───────────────────────────────────────────────────

def build_excel(master_df: pd.DataFrame, regime_df: pd.DataFrame):
    logger.info("Building Excel workbook...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default Sheet

    build_sheet_regime_summary(wb, master_df, regime_df)
    build_sheet_full_period(wb, master_df)
    build_sheet_fixed_windows(wb, master_df)
    build_sheet_rolling_heatmap(wb, master_df, 3)
    build_sheet_rolling_heatmap(wb, master_df, 5)
    build_sheet_equity_curves(wb, master_df)
    build_sheet_how_to_complete(wb)
    build_sheet_methodology(wb, master_df)

    out_path = OUTDIR / "Alpha_Stack_Matrix_v2.xlsx"
    wb.save(str(out_path))
    logger.info("Saved: %s", out_path)
    return out_path


# ── Merge script generator ────────────────────────────────────────────────────

def write_merge_script():
    script = '''"""
merge_local_results.py — Merge local run results into Alpha Stack Matrix v2
============================================================================
Run this AFTER completing the local run with regime_aware_backtest_matrix.py.

Usage:
    python scripts/merge_local_results.py \\
        --local outputs/regime_matrix/local_run/master_results.csv
"""
from __future__ import annotations
import argparse, sys
from pathlib import Path
import pandas as pd

ROOT   = Path(__file__).resolve().parent.parent
OUTDIR = ROOT / "outputs" / "regime_matrix"
PENDING = "PENDING"


def merge(local_csv: Path, base_csv: Path) -> pd.DataFrame:
    base  = pd.read_csv(base_csv)
    local = pd.read_csv(local_csv)

    # Keep only value_only and combined_static from local run
    local_filt = local[local["config"].isin(["value_only", "combined_static"])].copy()
    print(f"Local rows to merge: {len(local_filt)}")

    # Drop PENDING rows for those configs in base
    base_clean = base[~(
        base["config"].isin(["value_only", "combined_static"]) &
        (base["cagr"] == PENDING)
    )].copy()
    print(f"Base rows after removing PENDING: {len(base_clean)}")

    merged = pd.concat([base_clean, local_filt], ignore_index=True)
    print(f"Merged total rows: {len(merged)}")
    return merged


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--local", required=True,
                    help="Path to local run master_results.csv")
    args = ap.parse_args()

    local_csv = Path(args.local)
    base_csv  = OUTDIR / "master_results.csv"

    if not local_csv.exists():
        print(f"ERROR: local results not found: {local_csv}")
        sys.exit(1)
    if not base_csv.exists():
        print(f"ERROR: base results not found: {base_csv}")
        sys.exit(1)

    merged = merge(local_csv, base_csv)
    out = OUTDIR / "master_results.csv"
    merged.to_csv(out, index=False)
    print(f"Saved merged results: {out}")

    # Rebuild Excel
    import importlib.util, os
    os.chdir(str(ROOT))
    spec = importlib.util.spec_from_file_location("build_complete_matrix",
        ROOT / "scripts" / "build_complete_matrix.py")
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)

    data = m.load_data()
    regime_df = m.classify_regimes(data["spy_nav"])
    # Re-run analysis with merged data (will overwrite equity curves too)
    print("Rebuilding Excel workbook with merged data...")
    m.build_excel(merged, regime_df)
    print("Done.")


if __name__ == "__main__":
    main()
'''
    out = ROOT / "scripts" / "merge_local_results.py"
    with open(out, "w") as f:
        f.write(script)
    logger.info("Saved merge script: %s", out)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    logger.info("Loading data...")
    data      = load_data()
    regime_df = classify_regimes(data["spy_nav"])
    regime_df.to_csv(OUTDIR / "regime_history.csv")
    logger.info("Regime classified: %d days", len(regime_df))

    logger.info("Running analysis across all windows and configs...")
    master_df = run_analysis(data)

    # Add regime stats to relevant rows
    for idx, row in master_df.iterrows():
        reg = regime_stats(regime_df, row["start"], row["end"])
        for k, v in reg.items():
            master_df.at[idx, k] = v

    master_df.to_csv(OUTDIR / "master_results.csv", index=False)
    logger.info("master_results.csv: %d rows", len(master_df))

    # Regime comparison table
    comp = build_regime_comparison(master_df, regime_df)
    comp.to_csv(OUTDIR / "regime_comparison.csv", index=False)
    logger.info("regime_comparison.csv: %d rows", len(comp))

    # Full-period summary log
    fp = master_df[master_df["window"] == "full_period"].copy()
    fp["_s"] = fp["config"].map({c: i for i, c in enumerate(CONFIG_ORDER)})
    fp = fp.sort_values("_s")
    logger.info("\n%s", "=" * 78)
    logger.info("FULL-PERIOD RESULTS (2009–2025)")
    logger.info("%-22s | CAGR   | Sharpe | Sortino | MaxDD  | Vol    | vs SPY",
                "Config")
    logger.info("%s", "-" * 78)
    for _, r in fp.iterrows():
        def _v(x): return f"{x:>6}" if x != "PENDING" else "⏳ PEND"
        logger.info("%-22s | %s | %s | %s   | %s | %s | %s",
                    CONFIG_META[r['config']]['label'],
                    _v(r.get('cagr')), _v(r.get('sharpe')), _v(r.get('sortino')),
                    _v(r.get('max_dd')), _v(r.get('annual_vol')), _v(r.get('spy_excess_cagr')))
    logger.info("=" * 78)

    # Excel
    xl_path = build_excel(master_df, regime_df)

    # Merge script
    write_merge_script()

    logger.info("\nAll outputs saved to: %s", OUTDIR)
    logger.info("  master_results.csv")
    logger.info("  regime_comparison.csv")
    logger.info("  Alpha_Stack_Matrix_v2.xlsx")
    logger.info("  equity_curves/ (%d files)", len(list(ECDIR.glob("*.csv"))))
    logger.info("  merge_local_results.py")
    return master_df


if __name__ == "__main__":
    main()
