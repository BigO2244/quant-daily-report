"""
run_value_backtest_csv.py — Value-Only & Combined 50/50 backtest (sandbox-safe)
================================================================================
Runs entirely from CSV files — no yfinance, no pyarrow, no network.

Prerequisites:
    Run scripts/export_matrix_inputs.py on your Mac first to produce:
        alpha_stack_cache/csv_export/prices_matrix.csv
        alpha_stack_cache/csv_export/edgar_facts.csv

Then run this script in the sandbox (system Python 3.10):
    python scripts/run_value_backtest_csv.py

Produces:
    outputs/regime_matrix/value_nav_timeseries.csv   (NAV per window per config)
    outputs/regime_matrix/master_results.csv          (updated with real values)
    outputs/regime_matrix/Alpha_Stack_Matrix.xlsx     (rebuilt, all 6 configs)
    outputs/regime_matrix/regime_comparison.csv       (updated)
    Prints the 4-question analysis to stdout and saves ANALYSIS.md
"""
from __future__ import annotations

import logging
import sys
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("value_backtest")

ROOT       = Path(__file__).resolve().parent.parent
CSV_DIR    = ROOT / "alpha_stack_cache" / "csv_export"
RES_DIR    = ROOT / "outputs" / "research"
OUTDIR     = ROOT / "outputs" / "regime_matrix"
ECDIR      = OUTDIR / "equity_curves"
OUTDIR.mkdir(parents=True, exist_ok=True)
ECDIR.mkdir(parents=True, exist_ok=True)

# ── Constants ─────────────────────────────────────────────────────────────────
RISK_FREE      = 0.04
COMMISSION_BPS = 5.0
SLIPPAGE_BPS   = 5.0
INITIAL_EQUITY = 100_000.0
TOP_N_VALUE    = 30
TOP_N_TREND    = 20
REBAL_RULE     = "ME"   # month-end

FIXED_WINDOWS = [
    ("full_period",   "2008-01-01", "2026-03-07"),
    ("regime_gfc",    "2008-01-01", "2010-12-31"),
    ("regime_bull1",  "2011-01-01", "2015-12-31"),
    ("regime_bull2",  "2016-01-01", "2020-12-31"),
    ("regime_covid",  "2019-01-01", "2022-12-31"),
    ("regime_recent", "2021-01-01", "2026-03-07"),
    ("last_3yr",      "2023-03-07", "2026-03-07"),
    ("last_5yr",      "2021-03-07", "2026-03-07"),
]

CONFIG_META = {
    "trend_only":         {"label": "Trend-Only",         "color": "2E75B6"},
    "value_only":         {"label": "Value-Only",          "color": "E7824A"},
    "combined_static":    {"label": "Combined 50/50",      "color": "70AD47"},
    "combined_allocator": {"label": "Combined Allocator",  "color": "7030A0"},
    "combined_alloc_cb":  {"label": "Alloc + CB",          "color": "C00000"},
    "spy_benchmark":      {"label": "SPY Benchmark",       "color": "595959"},
}


# ── Local staging dir for FUSE-safe output ────────────────────────────────────
# Write to the csv_export dir first (always writable), then copy to
# outputs/regime_matrix to avoid FUSE EDEADLK on writes.
import shutil, tempfile as _tempfile
_TMP_OUT = CSV_DIR / "matrix_outputs"
_TMP_OUT.mkdir(parents=True, exist_ok=True)


def _safe_write_csv(df: pd.DataFrame, dest: Path, index: bool = False) -> None:
    """Write df to dest via /tmp to avoid FUSE mount write locks."""
    tmp = _TMP_OUT / dest.name
    df.to_csv(tmp, index=index)
    try:
        shutil.copy2(str(tmp), str(dest))
    except OSError as e:
        logger.warning("Could not copy %s to mount (%s); file is at %s", dest.name, e, tmp)


def _safe_write_bytes(data: bytes, dest: Path) -> None:
    """Write raw bytes to dest via /tmp."""
    tmp = _TMP_OUT / dest.name
    tmp.write_bytes(data)
    try:
        shutil.copy2(str(tmp), str(dest))
    except OSError as e:
        logger.warning("Could not copy %s to mount (%s); file is at %s", dest.name, e, tmp)


# ═══════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ═══════════════════════════════════════════════════════════════════════════════

def load_prices() -> pd.DataFrame:
    """Load wide price DataFrame (date × ticker) from CSV export."""
    p = CSV_DIR / "prices_matrix.csv"
    if not p.exists():
        logger.error("prices_matrix.csv not found. Run export_matrix_inputs.py on Mac first.")
        sys.exit(1)
    logger.info("Loading prices from CSV (%s)...", p.name)
    df = pd.read_csv(p, index_col=0, parse_dates=True)
    df.index.name = "date"
    logger.info("  Prices: %d dates × %d tickers (%s → %s)",
                len(df), len(df.columns),
                df.index[0].date(), df.index[-1].date())
    return df


def load_edgar_facts() -> pd.DataFrame:
    """Load combined EDGAR facts CSV."""
    p = CSV_DIR / "edgar_facts.csv"
    if not p.exists():
        logger.error("edgar_facts.csv not found. Run export_matrix_inputs.py on Mac first.")
        sys.exit(1)
    logger.info("Loading EDGAR facts from CSV (%s)...", p.name)
    df = pd.read_csv(p, parse_dates=["filed", "end"], low_memory=False)
    logger.info("  EDGAR: %d rows, %d tickers", len(df), df["ticker"].nunique())
    return df


def load_universe() -> pd.DataFrame:
    return pd.read_csv(ROOT / "data" / "universe.csv")


def load_existing_nav() -> tuple[pd.Series | None, pd.Series | None,
                                  pd.Series | None, pd.Series | None]:
    """
    Load the existing confirmed NAV series (trend, allocator, CB, SPY).
    Returns (None, None, None, None) gracefully if files are inaccessible
    (e.g. FUSE mount locks) — callers must handle None returns.
    """
    try:
        tr  = pd.read_csv(RES_DIR / "sleeve1_backtest_2009_2025_timeseries.csv",
                          parse_dates=["date"]).set_index("date").sort_index()
        av  = pd.read_csv(RES_DIR / "sleeve1_alpha_variant_timeseries.csv",
                          parse_dates=["date"]).set_index("date").sort_index()
        trend_nav    = tr["portfolio_nav"]
        spy_nav      = tr["spy_nav"]
        alloc_nav    = av["net_nav"]
        alloc_nav_cb = av["net_nav_cb"]
        return trend_nav, alloc_nav, alloc_nav_cb, spy_nav
    except OSError as e:
        logger.warning("Could not load existing NAV files (likely FUSE lock): %s", e)
        logger.warning("Trend weights will be built from price momentum directly.")
        return None, None, None, None


# ═══════════════════════════════════════════════════════════════════════════════
# VALUE SIGNAL ENGINE  (PIT-safe, vectorized, sector-relative)
# ═══════════════════════════════════════════════════════════════════════════════

FLOW_FIELDS  = {"operating_cf", "capex", "net_income"}          # use TTM (quarterly sum)
STOCK_FIELDS = {"equity", "eps_diluted"}                         # use latest point-in-time
VALUE_FIELDS = ["net_income", "operating_cf", "capex", "eps_diluted", "equity"]
_Q_FORMS     = {"10-Q", "10-Q/A"}
_A_FORMS     = {"10-K", "10-K/A"}
_ALL_FORMS   = _Q_FORMS | _A_FORMS


def _build_fact_lookup(facts: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """
    Pre-compute per-field lookup tables (filed_date × ticker → value).
    Returns dict: field_name → DataFrame(index=filed DatetimeIndex, columns=tickers).
    For flow fields (net_income, operating_cf, capex), computes rolling 4-quarter TTM.
    For stock fields, uses latest filed value.

    This runs ONCE and is O(N log N). Each per-date lookup then becomes
    a simple ffill + reindex operation — orders of magnitude faster than
    the previous per-ticker per-date linear scan.
    """
    logger.info("Pre-building EDGAR fact lookup tables (vectorized)...")
    sub = facts[facts["form"].isin(_ALL_FORMS)].copy()
    sub["filed"] = pd.to_datetime(sub["filed"], errors="coerce")
    sub["val"]   = pd.to_numeric(sub["val"],   errors="coerce")
    sub = sub.dropna(subset=["filed", "val"])
    sub = sub[sub["field_name"].isin(VALUE_FIELDS)]

    lookup: dict[str, pd.DataFrame] = {}

    for field in VALUE_FIELDS:
        f_sub = sub[sub["field_name"] == field].copy()
        if f_sub.empty:
            lookup[field] = pd.DataFrame()
            continue

        if field in FLOW_FIELDS:
            # ── TTM via rolling 4-quarter sum ────────────────────────────────
            q = (f_sub[f_sub["form"].isin(_Q_FORMS)]
                 .sort_values(["ticker", "filed"])
                 .drop_duplicates(["ticker", "filed"], keep="last"))
            frames = []
            for ticker, g in q.groupby("ticker", sort=False):
                s = g.set_index("filed")["val"].sort_index()
                s = s[~s.index.duplicated(keep="last")]
                ttm = s.rolling(4, min_periods=3).sum()
                ttm.name = ticker
                frames.append(ttm)
            if frames:
                lookup[field] = pd.concat(frames, axis=1).sort_index()
                continue
            # Fallback to annual if no quarterly data
            f_sub = f_sub[f_sub["form"].isin(_A_FORMS)]

        # ── Point-in-time: latest filed per (ticker, date) ───────────────────
        pt = (f_sub.sort_values(["ticker", "filed"])
              .drop_duplicates(["ticker", "filed"], keep="last"))
        frames = []
        for ticker, g in pt.groupby("ticker", sort=False):
            s = g.set_index("filed")["val"].sort_index()
            s = s[~s.index.duplicated(keep="last")]
            s.name = ticker
            frames.append(s)
        lookup[field] = pd.concat(frames, axis=1).sort_index() if frames else pd.DataFrame()

    logger.info("  Lookup tables built for %d fields", len(lookup))
    return lookup


def _pit_at_dates(lookup: dict[str, pd.DataFrame],
                  field: str,
                  dates: pd.DatetimeIndex) -> pd.DataFrame:
    """
    Return a DataFrame (index=dates, columns=tickers) with PIT-safe values.
    Forward-fills from the most recent filing on or before each date.
    """
    lkp = lookup.get(field)
    if lkp is None or lkp.empty:
        return pd.DataFrame(index=dates)
    # Expand to union of filing dates + target dates, forward fill, then select
    all_idx = lkp.index.union(dates).sort_values()
    expanded = lkp.reindex(all_idx).ffill()
    return expanded.reindex(dates)


def _sector_z(series: pd.Series, sector_labels: pd.Series) -> pd.Series:
    """Within-sector z-score; falls back to cross-sectional if sector has < 3 names."""
    result = pd.Series(np.nan, index=series.index)
    for sec in sector_labels.unique():
        mask = sector_labels == sec
        s = series[mask].dropna()
        if len(s) < 3:
            continue
        std = s.std()
        if std > 1e-9:
            result[mask & series.notna()] = (series[mask & series.notna()] - s.mean()) / std
    return result


def build_value_scores_panel(facts: pd.DataFrame,
                              prices: pd.DataFrame,
                              universe: pd.DataFrame,
                              rebalance_dates: pd.DatetimeIndex
                              ) -> pd.DataFrame:
    """
    Vectorized replacement for compute_value_scores called in a loop.
    Returns a DataFrame: index=(date, ticker), columns=composite_z (and sub-scores).
    Runs in seconds rather than minutes.
    """
    sector_map  = universe.set_index("ticker")["sector"].to_dict()
    all_tickers = [t for t in universe["ticker"].dropna() if t in prices.columns]

    # Pre-build fact lookup once
    lookup = _build_fact_lookup(facts)

    # Get PIT values at all rebalance dates for each required field
    dates = rebalance_dates

    def _align(df: pd.DataFrame) -> pd.DataFrame:
        """Reindex to all_tickers (silently adds NaN for missing tickers)."""
        return df.reindex(columns=all_tickers)

    equity_df = _align(_pit_at_dates(lookup, "equity",       dates))
    ni_df     = _align(_pit_at_dates(lookup, "net_income",   dates))
    ocf_df    = _align(_pit_at_dates(lookup, "operating_cf", dates))
    capex_df  = _align(_pit_at_dates(lookup, "capex",        dates))
    eps_df    = _align(_pit_at_dates(lookup, "eps_diluted",  dates))

    # Close prices at rebalance dates (forward-fill)
    all_px_idx = prices.index.union(dates).sort_values()
    px_at_rebal = _align(prices.reindex(all_px_idx).ffill().reindex(dates))

    # ── Market cap via derived shares ────────────────────────────────────────
    # shares_outstanding is not in EDGAR CSV; derive from net_income / eps_diluted.
    # When both are negative (loss), the ratio is positive (correct share count).
    # Mask near-zero EPS to avoid division noise.
    eps_safe = eps_df.where(eps_df.abs() > 0.05)           # mask tiny EPS
    derived_shares = ni_df / eps_safe                       # (dates × tickers)
    derived_shares = derived_shares.where(derived_shares > 1e4)  # must be >10k shares

    mktcap = derived_shares * px_at_rebal                  # (dates × tickers)
    mktcap = mktcap.where(mktcap > 0)                      # mask non-positive

    # ── Signal ratios ────────────────────────────────────────────────────────
    # Earnings yield: EPS/Price is dimensionally correct (no mktcap needed)
    ey_panel  = eps_df / px_at_rebal.where(px_at_rebal > 0)  # EPS / price = E/P ✓
    # FCF yield: (OCF - Capex) / Mktcap
    fcf_panel = (ocf_df - capex_df.abs()) / mktcap
    # Book-to-price: Equity / Mktcap
    bp_panel  = equity_df / mktcap

    # ── Sector-relative z-scores then composite ───────────────────────────────
    sectors_s = pd.Series({t: sector_map.get(t, "Unknown") for t in all_tickers})
    records = []

    for dt in dates:
        def _get_row(panel_df):
            if dt not in panel_df.index:
                return pd.Series(dtype=float)
            return panel_df.loc[dt]

        row_ey  = _get_row(ey_panel)
        row_fcf = _get_row(fcf_panel)
        row_bp  = _get_row(bp_panel)

        z_scores = {}
        for name, row in [("ey", row_ey), ("fcf", row_fcf), ("bp", row_bp)]:
            valid = row.notna()
            if valid.sum() < 5:
                z_scores[name] = pd.Series(dtype=float)
                continue
            z_scores[name] = _sector_z(row[valid], sectors_s[valid])

        # Combine into composite (mean of available z-scores)
        z_df = pd.DataFrame(z_scores)
        if z_df.empty or z_df.isna().all().all():
            continue
        composite = z_df.mean(axis=1, skipna=True).dropna()
        for ticker, score in composite.items():
            records.append({"date": dt, "ticker": ticker,
                             "composite_z": float(score),
                             "sector": sector_map.get(ticker, "Unknown")})

    panel = pd.DataFrame(records)
    n_dates = panel["date"].nunique() if not panel.empty else 0
    n_tickers = panel["ticker"].nunique() if not panel.empty else 0
    logger.info("  Value scores panel: %d rows, %d dates, %d tickers avg coverage",
                len(panel), n_dates,
                len(panel) // n_dates if n_dates else 0)
    return panel


# Keep _pit_latest for any legacy callers (not used in the hot path)
def _pit_latest(facts: pd.DataFrame, ticker: str, field: str,
                as_of: pd.Timestamp, ttm: bool = False) -> float | None:
    """Legacy point-in-time lookup. Use build_value_scores_panel for bulk operations."""
    sub = facts[(facts["ticker"] == ticker) &
                (facts["field_name"] == field) &
                (facts["filed"] <= as_of) &
                (facts["form"].isin(_ALL_FORMS))]
    if sub.empty:
        return None
    if ttm:
        q = (sub[sub["form"].isin(_Q_FORMS)]
             .sort_values("filed").drop_duplicates("end", keep="last").tail(4))
        if len(q) >= 3:
            return float(q["val"].sum())
        a = sub[sub["form"].isin(_A_FORMS)].sort_values("filed")
        return float(a.iloc[-1]["val"]) if not a.empty else None
    return float(sub.sort_values("filed").iloc[-1]["val"])


def compute_value_scores(facts: pd.DataFrame, prices: pd.DataFrame,
                         universe: pd.DataFrame, as_of: pd.Timestamp) -> pd.DataFrame:
    """Single-date wrapper kept for compatibility. Not used in the hot path."""
    panel = build_value_scores_panel(facts, prices, universe,
                                     pd.DatetimeIndex([as_of]))
    if panel.empty:
        return pd.DataFrame()
    return panel[panel["date"] == as_of].rename(columns={"composite_z": "composite_z"})


# ═══════════════════════════════════════════════════════════════════════════════
# BACKTEST ENGINE  (self-contained, no external dependency)
# ═══════════════════════════════════════════════════════════════════════════════

def _cost_fraction(old_w: pd.Series, new_w: pd.Series) -> float:
    """Total cost as fraction of portfolio value for a rebalance."""
    turnover = (new_w.sub(old_w, fill_value=0).abs().sum()) / 2
    return turnover * (COMMISSION_BPS + SLIPPAGE_BPS) / 10_000


def run_weights_backtest(
    weights_df: pd.DataFrame,   # (rebal_dates × tickers), each row sums ≤ 1
    prices: pd.DataFrame,       # (all dates × tickers)
    spy_prices: pd.Series,      # date-indexed SPY prices
    window_start: str,
    window_end: str,
) -> dict:
    """
    Run a weights-based backtest with T+1 execution lag, costs, and SPY benchmark.

    Returns dict with keys:
        nav            — daily NAV Series (starts at 1.0)
        gross_nav      — nav before costs
        spy_nav        — SPY normalised to 1.0 at start
        turnover       — annualised average 1-way turnover (monthly × 12)
        avg_holdings   — average number of positions held
        stats          — dict of performance metrics
    """
    ws = pd.Timestamp(window_start)
    we = pd.Timestamp(window_end)

    # Price returns on all dates in range
    px = prices[(prices.index >= ws) & (prices.index <= we)].copy()
    if px.empty:
        return {}

    # SPY returns
    spy = spy_prices[(spy_prices.index >= ws) & (spy_prices.index <= we)].dropna()
    spy = spy / spy.iloc[0] if not spy.empty else spy

    all_dates = px.index
    # Calculate daily returns
    rets = px.pct_change().fillna(0.0)

    # Weights scheduled on rebal dates within window
    rebal_in_window = weights_df.index[(weights_df.index >= ws) & (weights_df.index <= we)]

    if len(rebal_in_window) == 0:
        return {}

    # Build daily weight series with T+1 lag (weight from date t applies from t+1 onwards)
    current_w = pd.Series(0.0, index=px.columns)
    daily_w   = pd.DataFrame(0.0, index=all_dates, columns=px.columns)
    prev_rebal_w = current_w.copy()

    total_cost = 0.0
    cost_by_date = pd.Series(0.0, index=all_dates)

    rebal_set = set(rebal_in_window)
    for i, dt in enumerate(all_dates):
        if dt in rebal_set:
            new_w = weights_df.loc[dt].reindex(px.columns, fill_value=0.0)
            cost = _cost_fraction(prev_rebal_w, new_w)
            cost_by_date.iloc[i] = cost
            total_cost += cost
            current_w = new_w.copy()
            prev_rebal_w = new_w.copy()
        daily_w.loc[dt] = current_w

    # NAV simulation
    nav       = np.ones(len(all_dates))
    gross_nav = np.ones(len(all_dates))

    for i in range(1, len(all_dates)):
        w = daily_w.iloc[i - 1]          # weights in effect today
        r = rets.iloc[i]                  # returns today
        port_ret = float((w * r).sum())
        gross_nav[i] = gross_nav[i - 1] * (1 + port_ret)
        net_ret   = port_ret - cost_by_date.iloc[i]
        nav[i]    = nav[i - 1] * (1 + net_ret)

    nav_s       = pd.Series(nav,       index=all_dates, name="nav")
    gross_nav_s = pd.Series(gross_nav, index=all_dates, name="gross_nav")

    # Turnover: annualised
    avg_holdings = (daily_w > 0.001).sum(axis=1).mean()
    n_rebals     = max(len(rebal_in_window) - 1, 1)
    n_years      = (we - ws).days / 365.25
    # Monthly rebal → ×12 for annual
    ann_turnover = (total_cost * 10_000 / (COMMISSION_BPS + SLIPPAGE_BPS) * 2 / n_rebals * 12)

    stats = _compute_stats(nav_s, gross_nav_s, spy, n_years)
    stats["turnover_annual"] = round(ann_turnover, 1)
    stats["avg_holdings"]    = round(avg_holdings, 1)

    return {
        "nav": nav_s,
        "gross_nav": gross_nav_s,
        "spy_nav": spy,
        "stats": stats,
    }


def _compute_stats(nav: pd.Series, gross_nav: pd.Series,
                   spy: pd.Series, n_years: float) -> dict:
    d     = nav.pct_change().dropna()
    gd    = gross_nav.pct_change().dropna()
    if len(d) < 10 or n_years < 0.1:
        return {}

    total  = float(nav.iloc[-1] - 1)
    cagr   = float((1 + total) ** (1 / n_years) - 1)
    vol    = float(d.std() * np.sqrt(252))
    sharpe = float((cagr - RISK_FREE) / vol) if vol > 1e-10 else 0.0

    down   = d[d < RISK_FREE / 252]
    d_vol  = float(down.std() * np.sqrt(252)) if len(down) > 1 else vol
    sortino = float((cagr - RISK_FREE) / d_vol) if d_vol > 1e-10 else 0.0

    cum  = (1 + d).cumprod()
    hwm  = cum.cummax()
    dds  = (cum - hwm) / hwm
    mdd  = float(dds.min())
    mdd_dt = str(dds.idxmin().date()) if not dds.isna().all() else "N/A"
    calmar = round(cagr / abs(mdd), 3) if mdd < -1e-4 else None

    gross_total = float(gross_nav.iloc[-1] - 1)
    gross_cagr  = float((1 + gross_total) ** (1 / n_years) - 1)
    win_rate    = float((d > 0).mean() * 100)

    # SPY stats
    spy_cagr = 0.0
    if not spy.empty and spy.iloc[0] > 0:
        sp = spy.reindex(nav.index, method="ffill").dropna()
        if len(sp) > 10:
            sp = sp / sp.iloc[0]
            spy_cagr = float((sp.iloc[-1]) ** (1 / n_years) - 1)

    return {
        "cagr":          round(cagr  * 100, 2),
        "gross_cagr":    round(gross_cagr * 100, 2),
        "net_cagr":      round(cagr  * 100, 2),
        "annual_vol":    round(vol   * 100, 2),
        "sharpe":        round(sharpe, 3),
        "sortino":       round(sortino, 3),
        "max_dd":        round(mdd   * 100, 2),
        "max_dd_date":   mdd_dt,
        "total_return":  round(total * 100, 2),
        "calmar":        calmar,
        "win_rate":      round(win_rate, 1),
        "spy_excess_cagr": round((cagr - spy_cagr) * 100, 2),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# WEIGHT GENERATION
# ═══════════════════════════════════════════════════════════════════════════════

def get_rebal_dates(start: str, end: str, prices: pd.DataFrame) -> pd.DatetimeIndex:
    """Month-end rebalance dates within the window."""
    ws, we = pd.Timestamp(start), pd.Timestamp(end)
    idx = prices[(prices.index >= ws) & (prices.index <= we)].index
    if idx.empty:
        return pd.DatetimeIndex([])
    monthly = idx.to_series().resample("ME").last()
    return pd.DatetimeIndex(monthly.dropna().values)


def build_value_weights(facts: pd.DataFrame, prices: pd.DataFrame,
                        universe: pd.DataFrame, rebal_dates: pd.DatetimeIndex,
                        top_n: int = TOP_N_VALUE,
                        data_coverage_tracker: list | None = None) -> pd.DataFrame:
    """
    Build value weight matrix (rebal_dates × all_tickers), equal-weighted top_n.
    Uses vectorized build_value_scores_panel — runs in seconds, not minutes.
    """
    all_tickers = [t for t in universe["ticker"].dropna() if t in prices.columns]
    weights = pd.DataFrame(0.0, index=rebal_dates, columns=all_tickers)

    # Single vectorized call across ALL rebalance dates
    panel = build_value_scores_panel(facts, prices, universe, rebal_dates)
    if panel.empty:
        logger.warning("build_value_weights: no scores produced by panel")
        return weights

    for dt in rebal_dates:
        dt_scores = panel[panel["date"] == dt]
        if dt_scores.empty:
            continue
        dt_scores = dt_scores[dt_scores["ticker"].isin(all_tickers)]
        top = dt_scores.nlargest(top_n, "composite_z")["ticker"].tolist()
        if top:
            w = 1.0 / len(top)
            for t in top:
                weights.loc[dt, t] = w
            if data_coverage_tracker is not None:
                data_coverage_tracker.append(len(top) / top_n)

    return weights


def build_trend_weights_from_existing(trend_nav: pd.Series,
                                      prices: pd.DataFrame,
                                      universe: pd.DataFrame,
                                      rebal_dates: pd.DatetimeIndex,
                                      top_n: int = TOP_N_TREND) -> pd.DataFrame:
    """
    Approximate trend weights using momentum signals from price data.
    Selects top_n tickers by 12-1 month momentum, equal-weighted.
    """
    all_tickers = [t for t in universe["ticker"].dropna() if t in prices.columns]
    weights = pd.DataFrame(0.0, index=rebal_dates, columns=all_tickers)

    for dt in rebal_dates:
        lookback_252 = dt - pd.DateOffset(days=252)
        lookback_21  = dt - pd.DateOffset(days=21)
        px_hist = prices[(prices.index >= lookback_252) & (prices.index <= dt)]
        if len(px_hist) < 50:
            continue
        # 12-1 month momentum: return from 252 days ago to 21 days ago
        px_252 = px_hist.iloc[0]
        px_21  = prices.loc[prices.index <= lookback_21].iloc[-1] if (prices.index <= lookback_21).any() else None
        if px_21 is None:
            continue
        mom = ((px_21 / px_252) - 1).dropna()
        mom = mom[mom.index.isin(all_tickers)]
        mom = mom[mom > 0]   # only positive momentum
        top = mom.nlargest(top_n).index.tolist()
        if top:
            w = 1.0 / len(top)
            for t in top:
                weights.loc[dt, t] = w

    return weights


# ═══════════════════════════════════════════════════════════════════════════════
# METRICS & ROLLING WINDOWS
# ═══════════════════════════════════════════════════════════════════════════════

def rolling_windows(window_years: int, step_months: int,
                    first: str = "2008-01-01", last: str = "2026-03-07") -> list:
    wins = []
    cur  = pd.Timestamp(first)
    lim  = pd.Timestamp(last)
    while True:
        end = cur + pd.DateOffset(years=window_years) - pd.DateOffset(days=1)
        if end > lim:
            break
        tag = f"roll_{window_years}y_{cur.year}q{(cur.month-1)//3+1}"
        wins.append((tag, str(cur.date()), str(end.date())))
        cur += pd.DateOffset(months=step_months)
    return wins


def nav_to_metrics(nav: pd.Series, gross_nav: pd.Series | None,
                   spy_nav: pd.Series, config: str, window: str,
                   start: str, end: str,
                   extra_stats: dict | None = None) -> dict:
    ws, we = pd.Timestamp(start), pd.Timestamp(end)
    n = nav[(nav.index >= ws) & (nav.index <= we)].dropna()
    if len(n) < 20:
        return {}
    n = n / n.iloc[0]
    n_years = max(len(n) / 252, 0.01)
    d    = n.pct_change().dropna()
    vol  = float(d.std() * np.sqrt(252))
    total = float(n.iloc[-1] - 1)
    cagr  = float((1 + total) ** (1 / n_years) - 1)
    sharpe = float((cagr - RISK_FREE) / vol) if vol > 1e-10 else 0.0
    down   = d[d < RISK_FREE / 252]
    dvol   = float(down.std() * np.sqrt(252)) if len(down) > 1 else vol
    sortino = float((cagr - RISK_FREE) / dvol) if dvol > 1e-10 else 0.0
    cum  = (1 + d).cumprod(); hwm = cum.cummax(); dds = (cum - hwm) / hwm
    mdd  = float(dds.min())
    mdd_dt = str(dds.idxmin().date()) if not dds.isna().all() else "N/A"
    calmar = round(cagr / abs(mdd), 3) if mdd < -1e-4 else None
    wr   = float((d > 0).mean() * 100)

    # Gross CAGR
    gc = np.nan
    if gross_nav is not None:
        gn = gross_nav[(gross_nav.index >= ws) & (gross_nav.index <= we)].dropna()
        if len(gn) >= 5:
            gn = gn / gn.iloc[0]
            gc = float((gn.iloc[-1]) ** (1 / max(len(gn)/252, 0.01)) - 1) * 100

    # SPY excess
    spy_sl = spy_nav[(spy_nav.index >= ws) & (spy_nav.index <= we)].dropna()
    spy_excess = np.nan
    if not spy_sl.empty:
        sp = spy_sl / spy_sl.iloc[0]
        sc = float((sp.iloc[-1]) ** (1 / max(len(sp)/252, 0.01)) - 1)
        spy_excess = round((cagr - sc) * 100, 2)

    r = {
        "window":    window,
        "config":    config,
        "start":     start,
        "end":       end,
        "n_years":   round(n_years, 2),
        "n_days":    len(n),
        "cagr":      round(cagr  * 100, 2),
        "gross_cagr":round(gc, 2) if not np.isnan(gc) else np.nan,
        "net_cagr":  round(cagr  * 100, 2),
        "annual_vol":round(vol   * 100, 2),
        "sharpe":    round(sharpe, 3),
        "sortino":   round(sortino, 3),
        "max_dd":    round(mdd   * 100, 2),
        "max_dd_date": mdd_dt,
        "total_return": round(total * 100, 2),
        "calmar":    calmar,
        "win_rate":  round(wr, 1),
        "spy_excess_cagr": spy_excess,
    }
    if extra_stats:
        r.update(extra_stats)
    return r


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER
# ═══════════════════════════════════════════════════════════════════════════════

def _fill(hex_c): return PatternFill("solid", fgColor=hex_c)
def _font(bold=False, size=10, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)
def _border():
    s = Side(style="thin", color="C0C0C0")
    return Border(left=s, right=s, top=s, bottom=s)
def _wc(ws, row, col, val, bold=False, fill=None, align="center", size=10,
        font_color="000000", italic=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = _font(bold=bold, size=size, color=font_color, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if fill: c.fill = fill
    c.border    = _border()
    return c

GOOD  = _fill("E2EFDA"); WARN = _fill("FCE4D6"); NEUT = _fill("DDEBF7")
ALT   = _fill("F2F2F2"); HDR  = _fill("1F3864"); SUB  = _fill("2E75B6")

def _mfill(v, metric):
    try: v = float(v)
    except: return None
    if metric == "sharpe":
        return GOOD if v >= 1.0 else (NEUT if v >= 0.5 else (WARN if v < 0.2 else None))
    if metric == "cagr":
        return GOOD if v >= 15 else (NEUT if v >= 8 else (WARN if v < 0 else None))
    if metric == "max_dd":
        return GOOD if v >= -10 else (NEUT if v >= -25 else (WARN if v < -40 else None))
    return None


def build_excel(master_df: pd.DataFrame) -> Path:
    logger.info("Building Excel workbook...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    CONFIG_ORDER = ["trend_only", "value_only", "combined_static",
                    "combined_allocator", "combined_alloc_cb", "spy_benchmark"]

    # ── Sheet 1: Regime Summary ─────────────────────────────────────────────
    ws = wb.create_sheet("Regime Summary")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:T1")
    _wc(ws,1,1,"ALPHA STACK — REGIME-AWARE BACKTEST MATRIX (ALL 6 CONFIGS)",
        bold=True, fill=HDR, font_color="FFFFFF", size=13, align="left")
    ws.row_dimensions[1].height = 28

    metrics3 = ["CAGR (%)", "Sharpe", "Max DD (%)"]
    col = 3
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    _wc(ws,4,1,"Window",bold=True,fill=HDR,font_color="FFFFFF")
    _wc(ws,4,2,"Dominant Regime",bold=True,fill=HDR,font_color="FFFFFF")

    cstart = {}
    for cfg in CONFIG_ORDER:
        if cfg not in CONFIG_META:
            continue
        meta = CONFIG_META[cfg]
        cstart[cfg] = col
        ws.merge_cells(start_row=3,start_column=col,end_row=3,end_column=col+2)
        c = ws.cell(row=3, column=col, value=meta["label"])
        c.font = _font(bold=True, color="FFFFFF")
        c.fill = _fill(meta["color"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        for i,lbl in enumerate(metrics3):
            _wc(ws,4,col+i,lbl,bold=True,fill=_fill(meta["color"]),font_color="FFFFFF",size=9)
            ws.column_dimensions[get_column_letter(col+i)].width = 9
        col += 3

    ws.row_dimensions[3].height = 20; ws.row_dimensions[4].height = 20

    alt = False
    for ri,(wn,ws_,we_,*_) in enumerate(FIXED_WINDOWS, start=5):
        bg = ALT if alt else None; alt = not alt
        meta_label = wn.replace("_"," ").title()
        _wc(ws,ri,1,meta_label,bold=(wn=="full_period"),
            fill=HDR if wn=="full_period" else bg,
            font_color="FFFFFF" if wn=="full_period" else "000000",align="left")
        _wc(ws,ri,2,"",fill=bg)
        for cfg in CONFIG_ORDER:
            if cfg not in cstart or cfg not in CONFIG_META:
                continue
            c = cstart[cfg]
            sub = master_df[(master_df["window"]==wn)&(master_df["config"]==cfg)]
            fp  = wn == "full_period"
            fc  = "FFFFFF" if fp else "000000"
            fill_base = HDR if fp else bg
            if sub.empty:
                for i in range(3): _wc(ws,ri,c+i,"—",fill=fill_base,font_color=fc,size=9)
                continue
            r = sub.iloc[0]
            for offset,(mname,mkey) in enumerate(zip(metrics3,["cagr","sharpe","max_dd"])):
                v = r.get(mkey)
                try: v_num = float(v)
                except: v_num = None
                disp = round(v_num,2) if v_num is not None else "—"
                mf = _mfill(v_num, mkey) if v_num is not None else None
                use_fill = fill_base if fp else (mf or bg)
                _wc(ws,ri,c+offset,disp,fill=use_fill,font_color=fc,size=9,bold=fp)
        ws.row_dimensions[ri].height = 16

    ws.freeze_panes = "C5"

    # ── Sheet 2: Full-Period Detail ─────────────────────────────────────────
    ws2 = wb.create_sheet("Full-Period Detail")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:M1")
    _wc(ws2,1,1,"FULL-PERIOD COMPARISON (2008–2026)",bold=True,fill=HDR,
        font_color="FFFFFF",size=12,align="left")
    ws2.row_dimensions[1].height = 26

    metric_cols = [("cagr","CAGR (%)"),("gross_cagr","Gross CAGR"),("annual_vol","Ann. Vol (%)"),
                   ("sharpe","Sharpe"),("sortino","Sortino"),("max_dd","Max DD (%)"),
                   ("calmar","Calmar"),("win_rate","Win Rate (%)"),("turnover_annual","Turnover"),
                   ("spy_excess_cagr","vs SPY (%)"),("avg_holdings","Avg Holdings"),("n_years","Years")]
    ws2.column_dimensions["A"].width = 24
    _wc(ws2,3,1,"Strategy",bold=True,fill=SUB,font_color="FFFFFF")
    for ci,(mk,hdr) in enumerate(metric_cols,start=2):
        ws2.column_dimensions[get_column_letter(ci)].width = 13
        _wc(ws2,3,ci,hdr,bold=True,fill=SUB,font_color="FFFFFF",size=9)
    ws2.row_dimensions[3].height = 20

    fp = master_df[master_df["window"]=="full_period"].copy()
    fp["_s"] = fp["config"].map({c:i for i,c in enumerate(CONFIG_ORDER)})
    fp = fp.sort_values("_s")
    for ri,(_, row) in enumerate(fp.iterrows(), start=4):
        cfg  = row["config"]
        meta = CONFIG_META.get(cfg, {"label": cfg, "color": "888888"})
        _wc(ws2,ri,1,meta["label"],bold=True,fill=_fill(meta["color"]),
            font_color="FFFFFF",align="left")
        for ci,(mk,_) in enumerate(metric_cols,start=2):
            v = row.get(mk)
            try: v_n = float(v)
            except: v_n = None
            disp = round(v_n,2) if v_n is not None else "—"
            _wc(ws2,ri,ci,disp,fill=_mfill(v_n,mk),size=9)
        ws2.row_dimensions[ri].height = 18
    ws2.freeze_panes = "B4"

    # ── Sheet 3: All Windows Table ──────────────────────────────────────────
    ws3 = wb.create_sheet("All Windows")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:K1")
    _wc(ws3,1,1,"ALL WINDOWS — KEY METRICS",bold=True,fill=HDR,font_color="FFFFFF",size=11,align="left")

    hdrs3 = ["Window","Config","CAGR (%)","Gross CAGR","Vol (%)","Sharpe","Sortino",
             "Max DD (%)","Win Rate (%)","Turnover","vs SPY (%)"]
    widths3= [22,22,9,9,9,8,8,9,10,9,10]
    for ci,(h,w) in enumerate(zip(hdrs3,widths3),start=1):
        ws3.column_dimensions[get_column_letter(ci)].width = w
        _wc(ws3,3,ci,h,bold=True,fill=SUB,font_color="FFFFFF",size=9)
    ws3.row_dimensions[3].height = 18

    fixed_names = [w[0] for w in FIXED_WINDOWS]
    fixed_df = master_df[master_df["window"].isin(fixed_names)].copy()
    fixed_df["_ws"] = fixed_df["window"].map({w:i for i,w in enumerate(fixed_names)})
    fixed_df["_cs"] = fixed_df["config"].map({c:i for i,c in enumerate(CONFIG_ORDER)})
    fixed_df = fixed_df.sort_values(["_ws","_cs"])

    cur_win = None; alt = False; ri = 4
    for _,row in fixed_df.iterrows():
        if row["window"] != cur_win: cur_win=row["window"]; alt=not alt
        bg = ALT if alt else None
        meta = CONFIG_META.get(row["config"],{"label":row["config"],"color":"888888"})
        vals = [row["window"].replace("_"," ").title(),
                meta["label"],
                row.get("cagr"), row.get("gross_cagr"), row.get("annual_vol"),
                row.get("sharpe"), row.get("sortino"), row.get("max_dd"),
                row.get("win_rate"), row.get("turnover_annual"), row.get("spy_excess_cagr")]
        keys = [None,None,"cagr","gross_cagr","annual_vol","sharpe","sortino","max_dd","win_rate",None,"spy_excess_cagr"]
        for ci,(v,k) in enumerate(zip(vals,keys),start=1):
            try: v_n = float(v)
            except: v_n = None
            disp = round(v_n,2) if v_n is not None else ("—" if v is None or (isinstance(v,float) and np.isnan(v)) else v)
            fill = _mfill(v_n, k) if k and v_n is not None else bg
            _wc(ws3,ri,ci,disp,fill=fill,size=8,align="left" if ci<=2 else "right" if isinstance(disp,(int,float)) else "center")
        ws3.row_dimensions[ri].height = 14; ri += 1

    ws3.freeze_panes = "C4"

    # ── Sheet 4: Rolling Sharpe Heatmap ────────────────────────────────────
    for yr in [3, 5]:
        wsH = wb.create_sheet(f"Rolling {yr}Y Sharpe")
        wsH.sheet_view.showGridLines = False
        wsH.merge_cells("A1:I1")
        _wc(wsH,1,1,f"ROLLING {yr}-YEAR WINDOWS — SHARPE HEATMAP",
            bold=True,fill=HDR,font_color="FFFFFF",size=11,align="left")

        roll_df = master_df[master_df["window"].str.startswith(f"roll_{yr}y_")]
        wins_sorted = sorted(roll_df["window"].unique())
        avail_cfgs  = [c for c in CONFIG_ORDER if c in roll_df["config"].unique()]

        _wc(wsH,3,1,"Start",bold=True,fill=SUB,font_color="FFFFFF"); wsH.column_dimensions["A"].width=12
        _wc(wsH,3,2,"End",  bold=True,fill=SUB,font_color="FFFFFF"); wsH.column_dimensions["B"].width=12
        for ci,cfg in enumerate(avail_cfgs,start=3):
            m = CONFIG_META.get(cfg,{"label":cfg,"color":"888888"})
            _wc(wsH,3,ci,m["label"],bold=True,fill=_fill(m["color"]),font_color="FFFFFF",size=9)
            wsH.column_dimensions[get_column_letter(ci)].width = 16

        for ri2,wn in enumerate(wins_sorted,start=4):
            wd = roll_df[roll_df["window"]==wn]
            st = wd["start"].iloc[0] if not wd.empty else ""
            en = wd["end"].iloc[0]   if not wd.empty else ""
            _wc(wsH,ri2,1,st,size=8); _wc(wsH,ri2,2,en,size=8)
            for ci,cfg in enumerate(avail_cfgs,start=3):
                sub = wd[wd["config"]==cfg]
                v   = round(float(sub.iloc[0]["sharpe"]),2) if not sub.empty else "—"
                _wc(wsH,ri2,ci,v,fill=_mfill(v,"sharpe") if isinstance(v,float) else None,size=8)
            wsH.row_dimensions[ri2].height = 13

        # Stats rows
        sr = 4 + len(wins_sorted) + 2
        wsH.merge_cells(start_row=sr,start_column=1,end_row=sr,end_column=3+len(avail_cfgs))
        _wc(wsH,sr,1,"Summary Statistics",bold=True,fill=_fill("D6DCE4"),align="left")
        for ri3,cfg in enumerate(avail_cfgs,start=sr+1):
            sub = roll_df[roll_df["config"]==cfg].copy()
            sh  = pd.to_numeric(sub["sharpe"],errors="coerce").dropna()
            cg  = pd.to_numeric(sub["cagr"],  errors="coerce").dropna()
            m   = CONFIG_META.get(cfg,{"label":cfg,"color":"888888"})
            _wc(wsH,ri3,1,m["label"],bold=True,fill=_fill(m["color"]),font_color="FFFFFF",size=9)
            _wc(wsH,ri3,2,f"Avg: {sh.mean():.2f}" if len(sh) else "—",size=9)
            _wc(wsH,ri3,3,f"Min: {sh.min():.2f}  Max: {sh.max():.2f}" if len(sh) else "—",size=9,align="left")

        wsH.freeze_panes = "C4"

    # ── Sheet 5: Analysis ───────────────────────────────────────────────────
    wsA = wb.create_sheet("Analysis & Conclusions")
    wsA.sheet_view.showGridLines = False
    wsA.column_dimensions["A"].width = 100
    _wc(wsA,1,1,"ALPHA STACK — INVESTMENT ANALYSIS & CONCLUSIONS",
        bold=True,fill=HDR,font_color="FFFFFF",size=13,align="left")
    wsA.row_dimensions[1].height = 28

    analysis_text = _build_analysis_text(master_df)
    for ri,line in enumerate(analysis_text.split("\n"),start=3):
        cell = wsA.cell(row=ri,column=1,value=line)
        cell.font = _font(bold=line.startswith("##") or line.startswith("Q"),
                         size=10 if line.startswith("#") else 9)
        if line.startswith("##"):
            cell.fill = _fill("D6DCE4")
        elif line.startswith("Q"):
            cell.fill = NEUT
        wsA.row_dimensions[ri].height = 15

    out_name = "Alpha_Stack_Matrix.xlsx"
    tmp_out = _TMP_OUT / out_name
    wb.save(str(tmp_out))
    dest = OUTDIR / out_name
    try:
        shutil.copy2(str(tmp_out), str(dest))
        logger.info("Saved workbook: %s", dest)
    except OSError as e:
        logger.warning("Could not copy xlsx to mount (%s); file is at %s", e, tmp_out)
        dest = tmp_out
    return dest


# ═══════════════════════════════════════════════════════════════════════════════
# ANALYSIS TEXT GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════

def _build_analysis_text(master_df: pd.DataFrame) -> str:
    fp = master_df[master_df["window"] == "full_period"].copy()
    get = lambda cfg, col: fp[fp["config"]==cfg][col].values[0] if len(fp[fp["config"]==cfg]) else None

    def fv(v, fmt=".1f"):
        if v is None or (isinstance(v, float) and np.isnan(v)): return "N/A"
        try: return format(float(v), fmt)
        except: return str(v)

    t_cagr   = get("trend_only",     "cagr");    v_cagr   = get("value_only",     "cagr")
    c_cagr   = get("combined_static","cagr");    a_cagr   = get("combined_allocator","cagr")
    t_sh     = get("trend_only",     "sharpe");  v_sh     = get("value_only",     "sharpe")
    c_sh     = get("combined_static","sharpe");  a_sh     = get("combined_allocator","sharpe")
    t_dd     = get("trend_only",     "max_dd");  v_dd     = get("value_only",     "max_dd")
    c_dd     = get("combined_static","max_dd");  a_dd     = get("combined_allocator","max_dd")
    spy_sh   = get("spy_benchmark",  "sharpe");  spy_cagr = get("spy_benchmark",  "cagr")

    def verdict(q, cond, pos, neg, neutral=None):
        if cond is None: return f"Q{q}: {neutral or 'Insufficient data.'}"
        return f"Q{q}: {'✅ YES — ' + pos if cond else '⚠️ NO — ' + neg}"

    lines = [
        "## ANALYSIS: DOES VALUE ADD VALUE ACROSS REGIMES?",
        "",
        f"Full-period baseline (2008–2026):",
        f"  Trend-Only:        CAGR {fv(t_cagr)}%  |  Sharpe {fv(t_sh,'0.3f')}  |  MaxDD {fv(t_dd)}%",
        f"  Value-Only:        CAGR {fv(v_cagr)}%  |  Sharpe {fv(v_sh,'0.3f')}  |  MaxDD {fv(v_dd)}%",
        f"  Combined 50/50:    CAGR {fv(c_cagr)}%  |  Sharpe {fv(c_sh,'0.3f')}  |  MaxDD {fv(c_dd)}%",
        f"  Combined Allocator:CAGR {fv(a_cagr)}%  |  Sharpe {fv(a_sh,'0.3f')}  |  MaxDD {fv(a_dd)}%",
        f"  SPY Benchmark:     CAGR {fv(spy_cagr)}%  |  Sharpe {fv(spy_sh,'0.3f')}",
        "",
        "─────────────────────────────────────────────────────────────────────",
        "",
        "Q1: DOES VALUE PRODUCE STANDALONE ALPHA?",
    ]

    try:
        q1_cond = float(v_sh) > float(spy_sh) if v_sh and spy_sh else None
        q1_pos  = f"Value-Only Sharpe {fv(v_sh,'0.3f')} > SPY {fv(spy_sh,'0.3f')}. Standalone alpha confirmed."
        q1_neg  = f"Value-Only Sharpe {fv(v_sh,'0.3f')} ≤ SPY {fv(spy_sh,'0.3f')}. Insufficient standalone edge."
    except: q1_cond = None; q1_pos = q1_neg = ""
    lines += [verdict("1",q1_cond,q1_pos,q1_neg), ""]

    lines.append("Q2: DOES VALUE REDUCE DRAWDOWNS vs TREND?")
    try:
        q2_cond = float(v_dd) > float(t_dd) if v_dd and t_dd else None  # less negative = better
        q2_pos  = f"Value MaxDD {fv(v_dd)}% is shallower than Trend {fv(t_dd)}%. Value is a drawdown reducer."
        q2_neg  = f"Value MaxDD {fv(v_dd)}% is deeper than Trend {fv(t_dd)}%. Value does not reduce drawdowns."
    except: q2_cond = None; q2_pos = q2_neg = ""
    lines += [verdict("2",q2_cond,q2_pos,q2_neg), ""]

    lines.append("Q3: DOES COMBINED 50/50 IMPROVE SHARPE vs TREND-ONLY?")
    try:
        delta_sh = float(c_sh) - float(t_sh) if c_sh and t_sh else None
        q3_cond  = delta_sh > 0.05 if delta_sh is not None else None
        q3_pos   = f"Combined Sharpe {fv(c_sh,'0.3f')} vs Trend {fv(t_sh,'0.3f')} (Δ +{fv(delta_sh,'0.3f')}). Diversification benefit confirmed."
        q3_neg   = f"Combined Sharpe {fv(c_sh,'0.3f')} vs Trend {fv(t_sh,'0.3f')} (Δ {fv(delta_sh,'0.3f')}). No material improvement from blend."
    except: q3_cond = None; q3_pos = q3_neg = ""
    lines += [verdict("3",q3_cond,q3_pos,q3_neg), ""]

    lines.append("Q4: DOES THE ALLOCATOR STILL ADD VALUE ONCE VALUE IS INCLUDED?")
    try:
        delta_a = float(a_sh) - float(c_sh) if a_sh and c_sh else None
        q4_cond = delta_a > 0.05 if delta_a is not None else None
        q4_pos  = f"Allocator Sharpe {fv(a_sh,'0.3f')} > 50/50 Sharpe {fv(c_sh,'0.3f')} (Δ +{fv(delta_a,'0.3f')}). Dynamic allocation adds value."
        q4_neg  = f"Allocator Sharpe {fv(a_sh,'0.3f')} ≤ 50/50 Sharpe {fv(c_sh,'0.3f')} (Δ {fv(delta_a,'0.3f')}). Allocator adds no edge over static blend."
    except: q4_cond = None; q4_pos = q4_neg = ""
    lines += [verdict("4",q4_cond,q4_pos,q4_neg), ""]

    # Regime-specific analysis
    lines += ["─────────────────────────────────────────────────────────────────────",
              "REGIME-SPECIFIC OBSERVATIONS", ""]
    regime_windows = [("GFC (2008-2010)", "regime_gfc"), ("Bull 1 (2011-2015)", "regime_bull1"),
                      ("Bull 2 (2016-2020)", "regime_bull2"), ("COVID (2019-2022)", "regime_covid"),
                      ("Recent (2021-2026)", "regime_recent")]
    for label, wn in regime_windows:
        sub = master_df[master_df["window"]==wn]
        if sub.empty: continue
        t_s = sub[sub["config"]=="trend_only"]["sharpe"].values
        v_s = sub[sub["config"]=="value_only"]["sharpe"].values
        c_s = sub[sub["config"]=="combined_static"]["sharpe"].values
        ts  = fv(t_s[0],"0.3f") if len(t_s) else "N/A"
        vs  = fv(v_s[0],"0.3f") if len(v_s) else "N/A"
        cs  = fv(c_s[0],"0.3f") if len(c_s) else "N/A"
        lines.append(f"  {label}: Trend Sharpe {ts} | Value Sharpe {vs} | 50/50 Sharpe {cs}")
    lines.append("")

    # Rolling robustness
    lines += ["─────────────────────────────────────────────────────────────────────",
              "ROLLING WINDOW ROBUSTNESS (3-YEAR)", ""]
    r3 = master_df[master_df["window"].str.startswith("roll_3y_")]
    for cfg in ["trend_only","value_only","combined_static","combined_allocator"]:
        sh = pd.to_numeric(r3[r3["config"]==cfg]["sharpe"],errors="coerce").dropna()
        m  = CONFIG_META.get(cfg,{"label":cfg})
        if len(sh):
            lines.append(f"  {m['label']:22s}: Avg Sharpe {sh.mean():.3f} | Min {sh.min():.3f} | Max {sh.max():.3f} | "
                         f"% > 1.0: {(sh>1.0).mean()*100:.0f}%  (n={len(sh)})")

    lines.append("")
    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════════
# REGIME COMPARISON TABLE
# ═══════════════════════════════════════════════════════════════════════════════

def build_regime_comparison(master_df: pd.DataFrame) -> pd.DataFrame:
    CONFIG_ORDER = ["trend_only","value_only","combined_static",
                    "combined_allocator","combined_alloc_cb","spy_benchmark"]
    rows = []
    for wn,ws_,we_ in [(w[0],w[1],w[2]) for w in FIXED_WINDOWS]:
        row = {"window": wn, "start": ws_, "end": we_}
        for cfg in CONFIG_ORDER:
            sub = master_df[(master_df["window"]==wn)&(master_df["config"]==cfg)]
            if sub.empty: continue
            r = sub.iloc[0]
            row[f"{cfg}_cagr"]   = r.get("cagr")
            row[f"{cfg}_sharpe"] = r.get("sharpe")
            row[f"{cfg}_maxdd"]  = r.get("max_dd")
        rows.append(row)
    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def _load_existing_master(outdir: Path) -> pd.DataFrame:
    """
    Load existing master_results.csv. Falls back to extracting from
    Alpha_Stack_Matrix_v2.xlsx if the CSV is FUSE-locked (EDEADLK).
    Returns an empty DataFrame if neither is accessible.
    """
    csv_path = outdir / "master_results.csv"
    try:
        df = pd.read_csv(csv_path)
        logger.info("Loaded existing master_results.csv (%d rows)", len(df))
        return df
    except OSError as e:
        logger.warning("master_results.csv inaccessible (%s), extracting from v2 Excel...", e)

    # ── Fallback: extract from Alpha_Stack_Matrix_v2.xlsx ─────────────────────
    v2_path = outdir / "Alpha_Stack_Matrix_v2.xlsx"
    if not v2_path.exists():
        logger.warning("Alpha_Stack_Matrix_v2.xlsx not found; starting from empty master.")
        return pd.DataFrame()

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(v2_path), read_only=True, data_only=True)
        ws = wb["Regime Summary"]
        rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
        wb.close()
    except Exception as ex:
        logger.warning("Could not read v2 xlsx: %s", ex)
        return pd.DataFrame()

    # Window-name → (start, end) mapping
    WIN_DATES = {
        "Full Period — 17 years":    ("2008-01-01", "2026-03-07"),
        "GFC Recovery (2009-2010)":  ("2008-01-01", "2010-12-31"),
        "Bull 1 — Post-GFC Rally":   ("2011-01-01", "2015-12-31"),
        "Bull 2 — Pre-COVID":        ("2016-01-01", "2020-12-31"),
        "COVID Crash + Recovery":    ("2019-01-01", "2022-12-31"),
        "Recent (2021-present)":     ("2021-01-01", "2026-03-07"),
        "Last 3 Years":              ("2023-03-07", "2026-03-07"),
        "Last 5 Years":              ("2021-03-07", "2026-03-07"),
    }
    WIN_KEY = {
        "Full Period — 17 years":    "full_period",
        "GFC Recovery (2009-2010)":  "regime_gfc",
        "Bull 1 — Post-GFC Rally":   "regime_bull1",
        "Bull 2 — Pre-COVID":        "regime_bull2",
        "COVID Crash + Recovery":    "regime_covid",
        "Recent (2021-present)":     "regime_recent",
        "Last 3 Years":              "last_3yr",
        "Last 5 Years":              "last_5yr",
    }

    # Header row detected at row index 3 (0-based) with 'Window' in col 0
    # Columns layout: window, dominant_regime,
    #   trend: cagr, sharpe, maxdd
    #   value: cagr, sharpe, maxdd  ← zeros (placeholder)
    #   comb:  cagr, sharpe, maxdd  ← zeros
    #   alloc: cagr, sharpe, maxdd
    # We extract trend_only and combined_allocator (non-zero) rows only.
    extracted = []
    for row in rows:
        if not isinstance(row[0], str) or row[0] not in WIN_KEY:
            continue
        wn     = WIN_KEY[row[0]]
        wdates = WIN_DATES[row[0]]
        ws_d, we_d = wdates
        n_years = (pd.Timestamp(we_d) - pd.Timestamp(ws_d)).days / 365.25

        def _safe(v, scale=1):
            if v is None or v == "" or (isinstance(v, str) and not v.strip()):
                return None
            try:
                f = float(v) * scale
                return round(f, 4)
            except (ValueError, TypeError):
                return None

        # trend_only at col indices 2,3,4
        # v2 Excel stores CAGR/MaxDD as percentages (19.4 means 19.4%) — same units
        # as nav_to_metrics which uses cagr*100.  Do NOT divide by 100.
        t_cagr, t_sharpe, t_maxdd = _safe(row[2]), _safe(row[3]), _safe(row[4])
        if t_cagr is not None and t_cagr != 0.0:
            extracted.append({"window": wn, "config": "trend_only",
                               "start": ws_d, "end": we_d, "n_years": round(n_years, 2),
                               "cagr": t_cagr, "sharpe": t_sharpe, "max_dd": t_maxdd,
                               "gross_cagr": None, "annual_vol": None, "sortino": None,
                               "turnover_annual": None, "avg_holdings": None,
                               "value_data_coverage": 0.0})
        # combined_allocator at col indices 11,12,13
        a_cagr, a_sharpe, a_maxdd = _safe(row[11]), _safe(row[12]), _safe(row[13])
        if a_cagr is not None and a_cagr != 0.0:
            extracted.append({"window": wn, "config": "combined_allocator",
                               "start": ws_d, "end": we_d, "n_years": round(n_years, 2),
                               "cagr": a_cagr, "sharpe": a_sharpe, "max_dd": a_maxdd,
                               "gross_cagr": None, "annual_vol": None, "sortino": None,
                               "turnover_annual": None, "avg_holdings": None,
                               "value_data_coverage": None})

    df = pd.DataFrame(extracted)
    logger.info("Extracted %d rows from v2 Excel (trend_only + combined_allocator)", len(df))
    return df


def main():
    logger.info("=" * 70)
    logger.info("VALUE BACKTEST — CSV-BASED RUNNER (sandbox-safe)")
    logger.info("=" * 70)

    # ── Load data ─────────────────────────────────────────────────────────────
    prices  = load_prices()
    facts   = load_edgar_facts()
    universe = load_universe()
    trend_nav, alloc_nav, alloc_nav_cb, spy_nav_raw = load_existing_nav()

    # SPY: always prefer price CSV column; fall back to loaded series
    if "SPY" in prices.columns:
        spy_prices = prices["SPY"].dropna()
    elif spy_nav_raw is not None:
        spy_prices = spy_nav_raw
    else:
        # Should not happen since prices CSV always has SPY, but be safe
        logger.error("SPY prices unavailable — cannot continue")
        sys.exit(1)

    logger.info("Universe: %d tickers", len(universe))
    logger.info("EDGAR facts: %d rows, %d tickers",
                len(facts), facts["ticker"].nunique())

    # ── Build all windows ─────────────────────────────────────────────────────
    all_windows = list(FIXED_WINDOWS) + \
                  [(t,s,e) for t,s,e in rolling_windows(3, 6)] + \
                  [(t,s,e) for t,s,e in rolling_windows(5, 12)]
    logger.info("Total windows: %d", len(all_windows))

    # ── Pre-compute rebalance weights for Value ───────────────────────────────
    # We compute weights for the FULL period once, then slice per window
    logger.info("Computing value weights for full period...")
    rebal_dates_full = get_rebal_dates("2007-01-01", "2026-03-07", prices)

    data_coverage_tracker: list = []
    val_weights_full = build_value_weights(
        facts, prices, universe, rebal_dates_full,
        top_n=TOP_N_VALUE, data_coverage_tracker=data_coverage_tracker
    )
    coverage = np.mean(data_coverage_tracker) if data_coverage_tracker else 0.0
    logger.info("Value weights computed: %d rebal dates, avg coverage %.0f%%",
                len(val_weights_full), coverage * 100)

    # Trend weights from price momentum
    logger.info("Computing trend weights...")
    trend_weights_full = build_trend_weights_from_existing(
        trend_nav, prices, universe, rebal_dates_full
    )

    # ── Run all windows × configs ──────────────────────────────────────────
    all_results   = []
    nav_timeseries = {}

    for wn, ws_, we_ in all_windows:
        ws_dt, we_dt = pd.Timestamp(ws_), pd.Timestamp(we_)
        rebal_in_win = rebal_dates_full[(rebal_dates_full >= ws_dt) &
                                        (rebal_dates_full <= we_dt)]

        if len(rebal_in_win) < 2:
            continue

        # ── Value-only ────────────────────────────────────────────────────
        v_weights = val_weights_full.reindex(rebal_in_win)
        if v_weights.notna().any().any():
            v_bt = run_weights_backtest(v_weights, prices, spy_prices, ws_, we_)
            if v_bt:
                extra = {"turnover_annual": v_bt["stats"].get("turnover_annual"),
                         "avg_holdings":    v_bt["stats"].get("avg_holdings"),
                         "value_data_coverage": round(coverage * 100, 1)}
                row = nav_to_metrics(v_bt["nav"], v_bt["gross_nav"], spy_prices,
                                     "value_only", wn, ws_, we_, extra)
                if row:
                    all_results.append(row)
                    nav_timeseries.setdefault(wn, {})["value_only"] = v_bt["nav"]

        # ── Combined 50/50 ────────────────────────────────────────────────
        t_weights = trend_weights_full.reindex(rebal_in_win)
        combined_weights = (
            v_weights.reindex(columns=t_weights.columns).fillna(0) * 0.5 +
            t_weights.reindex(columns=v_weights.columns).fillna(0) * 0.5
        )
        # Normalise rows (may not sum to exactly 1 if trend picks ≠ value picks)
        row_sums = combined_weights.sum(axis=1).replace(0, np.nan)
        combined_weights = combined_weights.div(row_sums, axis=0).fillna(0)

        if combined_weights.notna().any().any():
            c_bt = run_weights_backtest(combined_weights, prices, spy_prices, ws_, we_)
            if c_bt:
                extra = {"turnover_annual": c_bt["stats"].get("turnover_annual"),
                         "avg_holdings":    c_bt["stats"].get("avg_holdings"),
                         "value_data_coverage": round(coverage * 100, 1)}
                row = nav_to_metrics(c_bt["nav"], c_bt["gross_nav"], spy_prices,
                                     "combined_static", wn, ws_, we_, extra)
                if row:
                    all_results.append(row)
                    nav_timeseries.setdefault(wn, {})["combined_static"] = c_bt["nav"]

    # ── SPY benchmark rows (computed directly from prices) ────────────────────
    logger.info("Computing SPY benchmark rows...")
    for wn, ws_, we_ in all_windows:
        spy_win = spy_prices[(spy_prices.index >= pd.Timestamp(ws_)) &
                             (spy_prices.index <= pd.Timestamp(we_))].dropna()
        if len(spy_win) < 20:
            continue
        spy_nav_norm = spy_win / spy_win.iloc[0]
        n_years = max(len(spy_nav_norm) / 252, 0.01)
        d = spy_nav_norm.pct_change().dropna()
        vol  = float(d.std() * np.sqrt(252))
        total = float(spy_nav_norm.iloc[-1] - 1)
        cagr  = float((1 + total) ** (1 / n_years) - 1)
        sharpe = float((cagr - RISK_FREE) / vol) if vol > 1e-10 else 0.0
        cum   = (1 + d).cumprod(); hwm = cum.cummax(); dds = (cum - hwm) / hwm
        mdd   = float(dds.min())
        all_results.append({
            "window": wn, "config": "spy_benchmark",
            "start": ws_, "end": we_,
            "n_years": round(n_years, 2), "n_days": len(spy_nav_norm),
            "cagr": round(cagr * 100, 2),
            "gross_cagr": round(cagr * 100, 2),
            "net_cagr": round(cagr * 100, 2),
            "annual_vol": round(vol * 100, 2),
            "sharpe": round(sharpe, 3),
            "sortino": None, "max_dd": round(mdd * 100, 2),
            "max_dd_date": None, "total_return": round(total * 100, 2),
            "turnover_annual": 0.0, "avg_holdings": 1,
            "sector_hhi": None, "sleeve_correlation": None,
            "allocator_defensive_frac": None, "value_data_coverage": None,
        })

    logger.info("New results: %d rows (value_only + combined_static + spy_benchmark)", len(all_results))

    # ── Merge with existing master_results ───────────────────────────────────
    existing = _load_existing_master(OUTDIR)
    # Drop old zero/placeholder rows for configs we're re-computing
    existing_clean = existing[
        ~existing["config"].isin(["value_only", "combined_static", "spy_benchmark"])
    ].copy()
    new_df = pd.DataFrame(all_results)
    master = pd.concat([existing_clean, new_df], ignore_index=True)

    # Sort: fixed windows first, then rolling, within window by config order
    CONFIG_ORDER = ["trend_only","value_only","combined_static",
                    "combined_allocator","combined_alloc_cb","spy_benchmark"]
    fixed_names  = [w[0] for w in FIXED_WINDOWS]
    master["_is_fixed"] = master["window"].isin(fixed_names).astype(int)
    master["_cs"]       = master["config"].map({c:i for i,c in enumerate(CONFIG_ORDER)}).fillna(99)
    master = master.sort_values(["_is_fixed", "window", "_cs"],
                                ascending=[False, True, True])
    master = master.drop(columns=["_is_fixed","_cs"]).reset_index(drop=True)

    _safe_write_csv(master, OUTDIR / "master_results.csv")
    logger.info("Updated master_results.csv: %d rows", len(master))

    # ── Save NAV timeseries ──────────────────────────────────────────────────
    nav_rows = []
    for wn, configs_nav in nav_timeseries.items():
        for cfg, nav_s in configs_nav.items():
            for dt, v in nav_s.items():
                nav_rows.append({"window": wn, "config": cfg, "date": dt, "nav": v})
    if nav_rows:
        _safe_write_csv(pd.DataFrame(nav_rows), OUTDIR / "value_nav_timeseries.csv", index=False)
        logger.info("Saved value_nav_timeseries.csv: %d rows", len(nav_rows))

    # ── Equity curves ────────────────────────────────────────────────────────
    for wn, configs_nav in nav_timeseries.items():
        ec_path = ECDIR / f"{wn}.csv"
        try:
            existing_ec = pd.read_csv(ec_path, index_col=0, parse_dates=True) \
                          if ec_path.exists() else pd.DataFrame()
        except OSError:
            existing_ec = pd.DataFrame()
        for cfg, nav_s in configs_nav.items():
            existing_ec[cfg] = nav_s
        _safe_write_csv(existing_ec, ec_path, index=True)

    # ── Regime comparison ────────────────────────────────────────────────────
    comp = build_regime_comparison(master)
    _safe_write_csv(comp, OUTDIR / "regime_comparison.csv", index=False)
    logger.info("Updated regime_comparison.csv")

    # ── Excel ─────────────────────────────────────────────────────────────────
    xl_path = build_excel(master)

    # ── Analysis text ─────────────────────────────────────────────────────────
    analysis = _build_analysis_text(master)
    analysis_content = "# Alpha Stack — Value-Sleeve Analysis\n\n" + analysis
    _safe_write_bytes(analysis_content.encode("utf-8"), OUTDIR / "ANALYSIS.md")

    # ── Print summary ──────────────────────────────────────────────────────────
    print("\n" + "=" * 70)
    print("FULL-PERIOD RESULTS")
    print("=" * 70)
    fp = master[master["window"] == "full_period"]
    fp["_s"] = fp["config"].map({c:i for i,c in enumerate(CONFIG_ORDER)}).fillna(99)
    for _,r in fp.sort_values("_s").iterrows():
        print(f"  {CONFIG_META.get(r['config'],{'label':r['config']})['label']:22s} | "
              f"CAGR {r.get('cagr','—'):>6} | Sharpe {r.get('sharpe','—'):>6} | "
              f"MaxDD {r.get('max_dd','—'):>7} | Vol {r.get('annual_vol','—'):>5}")

    print("\n" + analysis)
    print("\nOutputs saved to:", OUTDIR)
    return master


if __name__ == "__main__":
    main()
